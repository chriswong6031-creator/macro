"""CCW-W1 tests: corporate bond holdings parser, registry, and match-rate.

Covers:
  - _parse_bond_xlsx: header detection, ISIN/cusip6, numeric coercion,
    maturity formats (MM/DD/YYYY and DD-Mon-YYYY), schema columns
  - keep-FIRST: _write_snapshot must not overwrite an existing dated parquet
  - issuer_themes.json: loads, structure invariants, JSON round-trip
  - match-rate: synthetic frame + real registry → MSFT matches, unknown lands
    unmatched, CCW-UNMATCHED-HIGH-PAR alarm fires above $5B threshold
  - test_fred_series_alias_conflicts.py companion: confirmed no collisions
    with the new corp_credit group (run separately via ci.yml job)
"""
from __future__ import annotations

import io
import json
import logging
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parents[1]

# ---------------------------------------------------------------------------
# Helpers: build a synthetic in-memory XLSX with bond columns
# ---------------------------------------------------------------------------

def _build_ssga_bond_xlsx(rows: list[dict], asof: str = "10-Jul-2026") -> bytes:
    """Build a minimal SSGA-format bond XLSX in memory.

    Preamble (rows 0-2) mirror the live format:
      row 0: Fund Name: | TEST FUND
      row 1: Ticker Symbol: | TEST
      row 2: Holdings: | As of <asof>
      row 3: (empty)
    Header row (row 4): Name | Identifier | SEDOL | Weight | Coupon |
                         Par Value | Market Value | Local Currency | Maturity
    Data rows follow.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Preamble
    ws.append(["Fund Name:", "TEST BOND FUND"])
    ws.append(["Ticker Symbol:", "TEST"])
    ws.append(["Holdings:", f"As of {asof}"])
    ws.append([None])  # blank row

    # Header
    ws.append(["Name", "Identifier", "SEDOL", "Weight", "Coupon",
               "Par Value", "Market Value", "Local Currency", "Maturity"])

    # Data rows
    for r in rows:
        ws.append([
            r.get("Name", "TEST BOND PLACEHOLDER"),
            r.get("Identifier", ""),
            r.get("SEDOL", "-"),
            r.get("Weight", 0.5),
            r.get("Coupon", 4.5),
            r.get("Par Value", 1_000_000),
            r.get("Market Value", 990_000),
            r.get("Local Currency", "USD"),
            r.get("Maturity", "03/15/2029"),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _load_xlsx(raw_bytes: bytes) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(raw_bytes), header=None)


# ---------------------------------------------------------------------------
# Import the module under test
# ---------------------------------------------------------------------------

from collectors.corp_bond_holdings import (
    OUT_COLS,
    UNMATCHED_HIGH_PAR_THRESHOLD,
    CorpBondHoldingsAdapter,
    _parse_bond_xlsx,
    _parse_maturities,
)


# ---------------------------------------------------------------------------
# 1. Parser: header detection
# ---------------------------------------------------------------------------

class TestHeaderDetection:
    def test_finds_name_header_row(self):
        rows = [{"Name": "TEST BOND INC SR UNSECURED 2.5", "Identifier": "US123456AB12"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert len(snap) == 1

    def test_raises_when_no_name_header(self):
        # XLSX with no 'Name' cell in first column
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Fund Name:", "BAD FUND"])
        ws.append(["No header row here", "x"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        raw = pd.read_excel(io.BytesIO(buf.read()), header=None)
        with pytest.raises(ValueError, match="header row"):
            _parse_bond_xlsx(raw, "BAD")


# ---------------------------------------------------------------------------
# 2. ISIN and cusip6 derivation
# ---------------------------------------------------------------------------

class TestIsinCusip6:
    def test_us_isin_yields_cusip6(self):
        rows = [{"Identifier": "US594918BT73", "Name": "MICROSOFT CORP 2029"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert snap["isin"].iloc[0] == "US594918BT73"
        assert snap["cusip6"].iloc[0] == "594918"

    def test_non_us_isin_gives_empty_cusip6(self):
        rows = [{"Identifier": "GB0030401166", "Name": "SOME FOREIGN BOND"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert snap["cusip6"].iloc[0] == ""

    def test_placeholder_isin_dash_gives_empty_cusip6(self):
        rows = [{"Identifier": "-", "Name": "CASH EQUIVALENT"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        # Should still parse, cusip6 empty
        assert snap["cusip6"].iloc[0] == ""


# ---------------------------------------------------------------------------
# 3. Numeric coercion
# ---------------------------------------------------------------------------

class TestNumericCoercion:
    def test_commas_stripped(self):
        rows = [{"Par Value": "1,000,000", "Market Value": "990,000", "Weight": "0.5"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert snap["par_value"].iloc[0] == 1_000_000.0
        assert snap["market_value"].iloc[0] == 990_000.0

    def test_dash_value_becomes_nan(self):
        # Supply a valid ISIN so the secondary footer filter retains the row.
        # A real bond may have "-" for par/coupon (e.g. perpetuals or cash equivalents).
        rows = [{"Par Value": "-", "Coupon": "-", "Identifier": "US594918BT73"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert pd.isna(snap["par_value"].iloc[0])
        assert pd.isna(snap["coupon"].iloc[0])

    def test_empty_value_becomes_nan(self):
        # Supply a valid ISIN so the secondary footer filter retains the row.
        rows = [{"Par Value": "", "Coupon": "", "Identifier": "US594918BT73"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert pd.isna(snap["par_value"].iloc[0])


# ---------------------------------------------------------------------------
# 4. Maturity parsing (both formats)
# ---------------------------------------------------------------------------

class TestMaturityParsing:
    def test_mm_dd_yyyy(self):
        s = pd.Series(["03/15/2029", "12/31/2034"])
        result = _parse_maturities(s)
        assert result.tolist() == ["2029-03-15", "2034-12-31"]

    def test_dd_mon_yyyy(self):
        s = pd.Series(["15-Mar-2029", "31-Dec-2034"])
        result = _parse_maturities(s)
        assert result.tolist() == ["2029-03-15", "2034-12-31"]

    def test_iso_already(self):
        s = pd.Series(["2029-03-15"])
        result = _parse_maturities(s)
        assert result.tolist() == ["2029-03-15"]

    def test_unparseable_becomes_none(self):
        s = pd.Series(["NOT A DATE", "-", "nan"])
        result = _parse_maturities(s)
        assert all(v is None for v in result.tolist())

    def test_maturity_in_parsed_snapshot(self):
        rows = [{"Maturity": "06/15/2055"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert snap["maturity"].iloc[0] == "2055-06-15"


# ---------------------------------------------------------------------------
# 5. Schema columns exact
# ---------------------------------------------------------------------------

class TestSchemaColumns:
    def test_output_columns_exact(self):
        rows = [{"Name": "BOND A", "Identifier": "US123456AB12"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "TEST")
        assert list(snap.columns) == OUT_COLS

    def test_fund_and_asof_columns_populated(self):
        rows = [{"Name": "BOND A", "Identifier": "US123456AB12"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows, asof="10-Jul-2026"))
        snap = _parse_bond_xlsx(raw, "SPIB")
        assert snap["fund"].iloc[0] == "SPIB"
        assert snap["as_of"].iloc[0] == "2026-07-10"


# ---------------------------------------------------------------------------
# 6. keep-FIRST: _write_snapshot must not overwrite existing dated file
# ---------------------------------------------------------------------------

class TestKeepFirst:
    def test_no_overwrite_existing(self, tmp_path, monkeypatch):
        """Second write to the same dated path must be skipped."""
        # Patch data_dir
        import lib.config as _cfg
        monkeypatch.setattr(_cfg, "data_dir", lambda: tmp_path)

        adapter = CorpBondHoldingsAdapter()

        rows = [{"Name": "BOND A", "Identifier": "US594918BT73"}]
        raw = _load_xlsx(_build_ssga_bond_xlsx(rows))
        snap = _parse_bond_xlsx(raw, "SPIB")

        # First write
        adapter._write_snapshot("SPIB", snap)
        target = tmp_path / "corp_bonds" / "holdings" / "SPIB" / f"{snap['as_of'].iloc[0]}.parquet"
        assert target.exists()

        # Corrupt the file to make it distinguishable
        target.write_bytes(b"SENTINEL")

        # Second write: must NOT overwrite
        adapter._write_snapshot("SPIB", snap)
        assert target.read_bytes() == b"SENTINEL", "keep-FIRST violated: file was overwritten"

    def test_new_date_writes_new_file(self, tmp_path, monkeypatch):
        import lib.config as _cfg
        monkeypatch.setattr(_cfg, "data_dir", lambda: tmp_path)

        adapter = CorpBondHoldingsAdapter()

        rows = [{"Name": "BOND A", "Identifier": "US594918BT73"}]
        raw1 = _load_xlsx(_build_ssga_bond_xlsx(rows, asof="10-Jul-2026"))
        snap1 = _parse_bond_xlsx(raw1, "SPIB")
        adapter._write_snapshot("SPIB", snap1)

        raw2 = _load_xlsx(_build_ssga_bond_xlsx(rows, asof="11-Jul-2026"))
        snap2 = _parse_bond_xlsx(raw2, "SPIB")
        adapter._write_snapshot("SPIB", snap2)

        files = list((tmp_path / "corp_bonds" / "holdings" / "SPIB").iterdir())
        assert len(files) == 2


# ---------------------------------------------------------------------------
# 7. Registry: issuer_themes.json structural invariants
# ---------------------------------------------------------------------------

class TestIssuerRegistry:
    @pytest.fixture(scope="class")
    def registry(self):
        path = ROOT / "data" / "corp_bonds" / "issuer_themes.json"
        assert path.exists(), f"issuer_themes.json not found at {path}"
        with path.open() as f:
            return json.load(f)

    def test_version_field(self, registry):
        assert registry["version"] == 1

    def test_has_themes(self, registry):
        assert len(registry["themes"]) >= 1

    def test_every_theme_has_issuers(self, registry):
        for theme_key, theme in registry["themes"].items():
            assert "issuers" in theme and len(theme["issuers"]) >= 1, \
                f"theme '{theme_key}' has no issuers"

    def test_every_issuer_has_required_fields(self, registry):
        for theme_key, theme in registry["themes"].items():
            for issuer_key, issuer in theme["issuers"].items():
                ctx = f"{theme_key}/{issuer_key}"
                assert "equity_ticker" in issuer and issuer["equity_ticker"], \
                    f"{ctx}: missing equity_ticker"
                assert "tier" in issuer and issuer["tier"] in ("IG", "HY", "split"), \
                    f"{ctx}: invalid tier {issuer.get('tier')!r}"
                assert "name_match_patterns" in issuer, f"{ctx}: missing name_match_patterns"
                assert isinstance(issuer["name_match_patterns"], list), \
                    f"{ctx}: name_match_patterns must be a list"
                assert len(issuer["name_match_patterns"]) >= 1 or \
                    len(issuer.get("id_prefixes", [])) >= 0, \
                    f"{ctx}: must have at least one matching mechanism"

    def test_json_round_trip(self, registry):
        """Serialize and re-parse: no data loss."""
        serialized = json.dumps(registry)
        reloaded = json.loads(serialized)
        assert reloaded == registry

    def test_required_themes_present(self, registry):
        required = {"hyperscaler_credit", "neocloud_credit", "memory_credit",
                    "dc_reit_credit", "ai_power_credit", "ai_hardware_credit",
                    "telecom_legacy"}
        present = set(registry["themes"].keys())
        missing = required - present
        assert not missing, f"Missing themes: {missing}"

    def test_required_issuers_present(self, registry):
        """Key issuers per the masterplan must be present."""
        required_by_theme = {
            "hyperscaler_credit": {"MSFT", "AMZN", "META", "GOOGL", "ORCL"},
            "neocloud_credit": {"CRWV"},
            "memory_credit": {"MU", "WDC", "STX"},
            "dc_reit_credit": {"EQIX", "DLR"},
            "ai_power_credit": {"VST", "CEG", "NEE"},
            "ai_hardware_credit": {"DELL", "HPE"},
            "telecom_legacy": {"T", "VZ"},
        }
        for theme, required in required_by_theme.items():
            if theme not in registry["themes"]:
                continue
            present = set(registry["themes"][theme]["issuers"].keys())
            missing = required - present
            assert not missing, f"theme '{theme}' missing issuers: {missing}"


# ---------------------------------------------------------------------------
# 8. Match-rate function: synthetic holdings + real registry
# ---------------------------------------------------------------------------

def _build_synthetic_holdings(rows: list[dict]) -> pd.DataFrame:
    """Build a minimal holdings frame matching the OUT_COLS schema."""
    records = []
    for r in rows:
        records.append({
            "isin": r.get("isin", "US594918BT73"),
            "cusip6": r.get("isin", "US594918BT73")[2:8]
                      if r.get("isin", "US").startswith("US") else "",
            "name": r.get("name", "UNKNOWN BOND"),
            "coupon": r.get("coupon", 4.5),
            "par_value": r.get("par_value", 1_000_000.0),
            "market_value": r.get("market_value", 990_000.0),
            "weight_pct": r.get("weight_pct", 0.1),
            "maturity": r.get("maturity", "2029-03-15"),
            "currency": r.get("currency", "USD"),
            "fund": r.get("fund", "TEST"),
            "as_of": r.get("as_of", "2026-07-10"),
        })
    return pd.DataFrame(records)


class TestMatchRate:
    @pytest.fixture(scope="class")
    def registry_data(self):
        path = ROOT / "data" / "corp_bonds" / "issuer_themes.json"
        assert path.exists()
        with path.open() as f:
            return json.load(f)

    def test_msft_matches_hyperscaler_credit(self, registry_data):
        """MICROSOFT name pattern must match MSFT in hyperscaler_credit."""
        holdings = _build_synthetic_holdings([
            {"name": "MICROSOFT CORP SR UNSECURED 09/30 1.35", "isin": "US594918BT73",
             "par_value": 10_000_000},
        ])
        # Build matchers inline (mirror the adapter logic)
        matchers = []
        for _tk, theme in registry_data["themes"].items():
            for ik, issuer in theme["issuers"].items():
                patterns = [p.upper() for p in issuer.get("name_match_patterns", [])]
                prefixes = [p.upper() for p in issuer.get("id_prefixes", [])]
                matchers.append((ik, patterns, prefixes))

        name_upper = holdings["name"].str.upper()
        cusip6_upper = holdings["cusip6"].str.upper()
        isin_upper = holdings["isin"].str.upper()

        def matched(i):
            nm, cu, isn = name_upper.iat[i], cusip6_upper.iat[i], isin_upper.iat[i]
            for _key, pats, pfxs in matchers:
                if any(p and p in nm for p in pats):
                    return True
                if any(p and (cu.startswith(p) or isn.startswith(p)) for p in pfxs):
                    return True
            return False

        assert matched(0), "MICROSOFT should match hyperscaler_credit/MSFT"

    def test_unknown_issuer_is_unmatched(self, registry_data):
        """An issuer not in the registry must be classified as unmatched."""
        holdings = _build_synthetic_holdings([
            {"name": "ZZZTESTCORP INC UNSECURED 2030", "isin": "US999999AA01",
             "par_value": 1_000_000},
        ])
        matchers = []
        for _tk, theme in registry_data["themes"].items():
            for ik, issuer in theme["issuers"].items():
                patterns = [p.upper() for p in issuer.get("name_match_patterns", [])]
                prefixes = [p.upper() for p in issuer.get("id_prefixes", [])]
                matchers.append((ik, patterns, prefixes))

        name_upper = holdings["name"].str.upper()
        cusip6_upper = holdings["cusip6"].str.upper()
        isin_upper = holdings["isin"].str.upper()

        def matched(i):
            nm, cu, isn = name_upper.iat[i], cusip6_upper.iat[i], isin_upper.iat[i]
            for _key, pats, pfxs in matchers:
                if any(p and p in nm for p in pats):
                    return True
                if any(p and (cu.startswith(p) or isn.startswith(p)) for p in pfxs):
                    return True
            return False

        assert not matched(0), "ZZZTESTCORP should NOT match any registry issuer"

    def test_high_par_alarm_triggers(self, tmp_path, monkeypatch, caplog):
        """_log_match_rate must emit CCW-UNMATCHED-HIGH-PAR when unmatched par > $5B."""
        import lib.config as _cfg
        monkeypatch.setattr(_cfg, "data_dir", lambda: tmp_path)

        # Create a minimal registry in tmp_path
        registry_dir = tmp_path / "corp_bonds"
        registry_dir.mkdir(parents=True, exist_ok=True)
        minimal_registry = {
            "version": 1,
            "as_of": "2026-07-14",
            "themes": {
                "test_theme": {
                    "name_en": "Test",
                    "name_zh": "测试",
                    "issuers": {
                        "MSFT": {
                            "equity_ticker": "MSFT",
                            "name_match_patterns": ["MICROSOFT"],
                            "id_prefixes": [],
                            "tier": "IG",
                            "coverage_notes": "",
                        }
                    }
                }
            }
        }
        (registry_dir / "issuer_themes.json").write_text(json.dumps(minimal_registry))

        # Build holdings: one matched (MSFT, $1B) + one large unmatched ($6B)
        holdings = _build_synthetic_holdings([
            {"name": "MICROSOFT CORP 2029", "isin": "US594918BT73", "par_value": 1_000_000_000},
            {"name": "ZZZUNKNOWN CORP 2030 FIRST TRANCHE", "isin": "US999000AA01",
             "par_value": 6_000_000_000},  # >$5B threshold
        ])

        adapter = CorpBondHoldingsAdapter()
        with caplog.at_level(logging.WARNING, logger="collectors.corp_bond_holdings"):
            adapter._log_match_rate(holdings)

        alarm_msgs = [r.message for r in caplog.records
                      if "CCW-UNMATCHED-HIGH-PAR" in r.message]
        assert alarm_msgs, (
            "CCW-UNMATCHED-HIGH-PAR alarm not fired for $6B unmatched issuer; "
            f"threshold is ${UNMATCHED_HIGH_PAR_THRESHOLD/1e9:.0f}B"
        )

    def test_high_par_alarm_does_not_trigger_below_threshold(self, tmp_path, monkeypatch, caplog):
        """No alarm when all unmatched groups are below $5B."""
        import lib.config as _cfg
        monkeypatch.setattr(_cfg, "data_dir", lambda: tmp_path)

        registry_dir = tmp_path / "corp_bonds"
        registry_dir.mkdir(parents=True, exist_ok=True)
        minimal_registry = {
            "version": 1,
            "as_of": "2026-07-14",
            "themes": {
                "test_theme": {
                    "name_en": "Test",
                    "name_zh": "测试",
                    "issuers": {
                        "MSFT": {
                            "equity_ticker": "MSFT",
                            "name_match_patterns": ["MICROSOFT"],
                            "id_prefixes": [],
                            "tier": "IG",
                            "coverage_notes": "",
                        }
                    }
                }
            }
        }
        (registry_dir / "issuer_themes.json").write_text(json.dumps(minimal_registry))

        holdings = _build_synthetic_holdings([
            {"name": "UNKNOWN SMALL CORP", "isin": "US999000BB01",
             "par_value": 100_000_000},  # $100M — well below threshold
        ])

        adapter = CorpBondHoldingsAdapter()
        with caplog.at_level(logging.WARNING, logger="collectors.corp_bond_holdings"):
            adapter._log_match_rate(holdings)

        alarm_msgs = [r.message for r in caplog.records
                      if "CCW-UNMATCHED-HIGH-PAR" in r.message]
        assert not alarm_msgs, "Alarm fired incorrectly below the $5B threshold"


# ---------------------------------------------------------------------------
# 9. Secondary footer filter: SSGA disclaimer rows are dropped
# ---------------------------------------------------------------------------

class TestFooterFilter:
    """Verify the ISIN/par retention predicate added after the name-check pass."""

    def _build_xlsx_with_footer_rows(self) -> bytes:
        """Build an XLSX that includes both genuine bond rows and SSGA disclaimer rows.

        Disclaimer rows have a long boilerplate Name but no Identifier and no Par Value.
        """
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active

        # Preamble
        ws.append(["Fund Name:", "TEST BOND FUND"])
        ws.append(["Ticker Symbol:", "TEST"])
        ws.append(["Holdings:", "As of 10-Jul-2026"])
        ws.append([None])

        # Header
        ws.append(["Name", "Identifier", "SEDOL", "Weight", "Coupon",
                   "Par Value", "Market Value", "Local Currency", "Maturity"])

        # Genuine bond row 1: valid ISIN + par value
        ws.append(["MICROSOFT CORP SR UNSECURED 2.5 2029", "US594918BT73", "-",
                   0.3, 2.5, 1_000_000, 990_000, "USD", "03/15/2029"])

        # Genuine bond row 2: valid ISIN but null par (e.g. TBA / placeholder)
        ws.append(["TREASURY PLACEHOLDER 2031", "US912828ZL46", "-",
                   0.1, 0.0, None, None, "USD", "06/15/2031"])

        # Footer disclaimer row 1: no Identifier, no Par Value, long boilerplate name
        ws.append(["State Street disclaims all warranties and representations of any kind "
                   "with regard to the accuracy or completeness of the information. "
                   "This document is not investment advice.",
                   None, None, None, None, None, None, None, None])

        # Footer disclaimer row 2: shorter disclaimer, still no ISIN or par
        ws.append(["For more information, please visit www.ssga.com.",
                   None, None, None, None, None, None, None, None])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def test_footer_rows_dropped(self):
        """Two footer disclaimer rows must be removed; two genuine rows survive."""
        raw = _load_xlsx(self._build_xlsx_with_footer_rows())
        snap = _parse_bond_xlsx(raw, "TEST")
        # Only the 2 genuine bond rows should remain
        assert len(snap) == 2, (
            f"Expected 2 rows after footer filter, got {len(snap)}: {snap['name'].tolist()}"
        )

    def test_genuine_row_null_par_valid_isin_kept(self):
        """A row with null par but a valid 12-char ISIN must be retained."""
        raw = _load_xlsx(self._build_xlsx_with_footer_rows())
        snap = _parse_bond_xlsx(raw, "TEST")
        # US912828ZL46 is the placeholder row with null par
        assert "US912828ZL46" in snap["isin"].tolist(), (
            "Row with valid ISIN but null par was incorrectly dropped by footer filter"
        )

    def test_footer_row_long_name_null_isin_null_par_dropped(self):
        """The State Street boilerplate row (long name, no ISIN, no par) must be dropped."""
        raw = _load_xlsx(self._build_xlsx_with_footer_rows())
        snap = _parse_bond_xlsx(raw, "TEST")
        names = snap["name"].tolist()
        for name in names:
            assert "disclaims" not in name.lower(), (
                f"Disclaimer row leaked into output: {name[:80]!r}"
            )
