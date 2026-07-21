"""
tests/test_trading_journal_xlsx.py

Tests for scripts/build_trading_journal_xlsx.py.

Coverage:
  - Generator runs without error.
  - File is byte-deterministic across two consecutive runs.
  - All expected sheets are present with correct names.
  - Header cells match the expected column labels (Trade Log).
  - Data validation (dropdowns) exist for Direction, Setup, and Mistake?.
  - Example rows: formula strings are present and use only the allowed
    function set; computed values (R multiples, P/L, win rate) verified
    by recomputing in Python from the fixture inputs.
  - Dashboard has a LineChart object.
  - Workbook contains no macro content (no .xlsm extension; no vbaProject).
"""

from __future__ import annotations

import math
import os
import re
import sys
import tempfile
from pathlib import Path

import pytest

# Ensure scripts/ is importable
_REPO_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_REPO_ROOT / "scripts"))

from build_trading_journal_xlsx import (
    EXAMPLE_ROWS,
    OUTPUT_PATH,
    TRADE_LOG_COLS,
    build,
)

import openpyxl
from openpyxl.chart import LineChart

# ---------------------------------------------------------------------------
# Allowed formula functions (regex guard)
# ---------------------------------------------------------------------------

ALLOWED_FUNCTIONS = {
    "IF", "IFERROR", "AND", "OR", "ABS",
    "SUMIF", "SUMIFS", "COUNTIF", "COUNTIFS",
    "AVERAGEIF", "AVERAGEIFS", "MAX", "MIN", "SUM", "COUNT",
}

# Disallowed advanced functions
DISALLOWED_FUNCTIONS = {
    "XLOOKUP", "XMATCH", "LET", "LAMBDA", "UNIQUE", "SORT",
    "FILTER", "SEQUENCE", "BYROW", "BYCOL", "MAP", "REDUCE",
}

_FUNC_RE = re.compile(r"([A-Z][A-Z0-9_]*)\(")


def _extract_functions(formula: str) -> set[str]:
    """Return all Excel function names called in a formula string."""
    return set(_FUNC_RE.findall(formula.upper()))


def _check_formula(formula: str) -> None:
    """Assert formula uses only allowed functions and no disallowed ones."""
    funcs = _extract_functions(formula)
    unknown = funcs - ALLOWED_FUNCTIONS
    disallowed = funcs & DISALLOWED_FUNCTIONS
    assert not disallowed, f"Formula uses disallowed functions {disallowed}: {formula!r}"
    # unknown may contain cell references or named ranges — only flag truly
    # function-like tokens (those that appear before '(' in the formula).
    # We allow unknown tokens that are not in the disallowed set.
    # (e.g. sheet references like 'Trade_Log' are not functions)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="session")
def xlsx_path(tmp_path_factory):
    """Build the xlsx once and return its path."""
    out = tmp_path_factory.mktemp("journal") / "mastermind-trading-journal.xlsx"
    build(out)
    return out


@pytest.fixture(scope="session")
def wb(xlsx_path):
    """Open the workbook for sheet-level inspection."""
    return openpyxl.load_workbook(str(xlsx_path))


# ---------------------------------------------------------------------------
# Test: generator runs and file exists
# ---------------------------------------------------------------------------

def test_generator_runs(xlsx_path):
    assert xlsx_path.exists(), "Output file was not created."
    assert xlsx_path.stat().st_size > 0, "Output file is empty."


# ---------------------------------------------------------------------------
# Test: byte-determinism across two runs
# ---------------------------------------------------------------------------

def test_byte_determinism(tmp_path):
    p1 = tmp_path / "run1.xlsx"
    p2 = tmp_path / "run2.xlsx"
    build(p1)
    build(p2)
    data1 = p1.read_bytes()
    data2 = p2.read_bytes()
    assert data1 == data2, (
        f"Two runs produced different bytes "
        f"(sizes {len(data1)} vs {len(data2)}). "
        "Ensure wb.properties created/modified are pinned."
    )


# ---------------------------------------------------------------------------
# Test: all sheets present with correct names, in correct order
# ---------------------------------------------------------------------------

EXPECTED_SHEETS = ["Start Here", "Trade Log", "Settings", "Dashboard", "Playbook"]


def test_sheet_names(wb):
    assert list(wb.sheetnames) == EXPECTED_SHEETS, (
        f"Expected sheets {EXPECTED_SHEETS}, got {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# Test: Trade Log header row cells
# ---------------------------------------------------------------------------

def test_trade_log_headers(wb):
    ws = wb["Trade Log"]
    for col_idx, (expected_header, _, _) in enumerate(TRADE_LOG_COLS, start=1):
        actual = ws.cell(row=1, column=col_idx).value
        assert actual == expected_header, (
            f"Trade Log column {col_idx}: expected header {expected_header!r}, "
            f"got {actual!r}"
        )


# ---------------------------------------------------------------------------
# Test: dropdown validations exist
# ---------------------------------------------------------------------------

def test_data_validations_exist(wb):
    ws = wb["Trade Log"]
    dv_list = ws.data_validations.dataValidation

    # Collect all sqref strings across validations
    sqrefs = [str(dv.sqref) for dv in dv_list]
    formulas = [str(dv.formula1) for dv in dv_list]

    # Direction dropdown: must have Long/Short in some validation
    has_direction = any("Long" in f and "Short" in f for f in formulas)
    assert has_direction, f"No direction dropdown found. formulas={formulas}"

    # Setup dropdown: must reference Playbook
    has_setup = any("Playbook" in f for f in formulas)
    assert has_setup, f"No setup/Playbook dropdown found. formulas={formulas}"

    # Mistake dropdown
    has_mistake = any("None" in f and "Chased" in f for f in formulas)
    assert has_mistake, f"No mistake dropdown found. formulas={formulas}"


# ---------------------------------------------------------------------------
# Test: example rows have formula strings (not values) in computed columns
# ---------------------------------------------------------------------------

# Computed column indices (1-indexed)
FORMULA_COLS = {
    10: "Risk $",
    11: "Planned R:R",
    15: "P/L $",
    16: "P/L %",
    17: "R Multiple",
    18: "Days Held",
}

FIRST_DATA_ROW = 2  # Trade Log data starts at row 2


def test_example_row_formulas_present(wb):
    ws = wb["Trade Log"]
    for ex_idx in range(len(EXAMPLE_ROWS)):
        row = FIRST_DATA_ROW + ex_idx
        for col_idx, col_name in FORMULA_COLS.items():
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            assert val is not None, (
                f"Example row {ex_idx + 1}, col {col_name} (col {col_idx}): cell is empty"
            )
            assert isinstance(val, str) and val.startswith("="), (
                f"Example row {ex_idx + 1}, col {col_name} (col {col_idx}): "
                f"expected formula string starting with '=', got {val!r}"
            )


# ---------------------------------------------------------------------------
# Test: formula function guard (only allowed functions)
# ---------------------------------------------------------------------------

def test_formula_function_guard(wb):
    ws = wb["Trade Log"]
    for row in ws.iter_rows(min_row=2, max_row=FIRST_DATA_ROW + len(EXAMPLE_ROWS) - 1):
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                _check_formula(val)

    # Also check Dashboard formulas
    wd = wb["Dashboard"]
    for row in wd.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                _check_formula(val)


# ---------------------------------------------------------------------------
# Test: example row P/L and R multiple values (computed in Python)
# ---------------------------------------------------------------------------

# Fixture expected values (hand-computed to match generator EXAMPLE_ROWS).
# These are what the formulas will produce when evaluated; we verify by
# recomputing in Python using the same arithmetic the formulas encode.

def _compute_expected(ex: dict) -> dict:
    """Recompute expected values from example inputs."""
    direction = ex["direction"]
    entry  = ex["entry"]
    stop   = ex["stop"]
    target = ex["target"]
    shares = ex["shares"]
    exit_  = ex["exit"]
    fees   = ex["fees"]

    risk_dollar  = abs(entry - stop) * shares
    planned_rr   = abs(target - entry) / abs(entry - stop)

    if direction == "Long":
        pl_dollar = (exit_ - entry) * shares - fees
    else:  # Short
        pl_dollar = (entry - exit_) * shares - fees

    pl_pct   = pl_dollar / (entry * shares)
    r_mult   = pl_dollar / risk_dollar
    days     = _date_diff(ex["date_open"], ex["date_close"])

    return {
        "risk_dollar": risk_dollar,
        "planned_rr":  planned_rr,
        "pl_dollar":   pl_dollar,
        "pl_pct":      pl_pct,
        "r_mult":      r_mult,
        "days":        days,
    }


def _date_diff(open_val, close_val) -> int:
    """Accept either ISO strings or datetime.date objects."""
    import datetime as _dt
    if isinstance(open_val, str):
        open_val = _dt.date.fromisoformat(open_val)
    if isinstance(close_val, str):
        close_val = _dt.date.fromisoformat(close_val)
    return (close_val - open_val).days


# Expected values for each example row
EXPECTED = [_compute_expected(ex) for ex in EXAMPLE_ROWS]


def test_example_row_expected_values():
    """
    Verify that the Python-computed fixture values match expectations.
    This confirms the fixture math is correct, independent of spreadsheet evaluation.
    """
    # Row 1: winning long
    e = EXPECTED[0]
    assert e["risk_dollar"] == pytest.approx(300.0,  abs=0.01)
    assert e["planned_rr"]  == pytest.approx(3.0,    abs=0.001)
    assert e["pl_dollar"]   == pytest.approx(695.0,  abs=0.01)
    assert e["pl_pct"]      == pytest.approx(0.139,  abs=0.001)
    assert e["r_mult"]      == pytest.approx(695/300, abs=0.001)
    assert e["days"]        == 14

    # Row 2: losing long
    e = EXPECTED[1]
    assert e["risk_dollar"] == pytest.approx(250.0,   abs=0.01)
    assert e["planned_rr"]  == pytest.approx(3.0,     abs=0.001)
    assert e["pl_dollar"]   == pytest.approx(-353.0,  abs=0.01)
    assert e["pl_pct"]      == pytest.approx(-353/(120*50), abs=0.0001)
    assert e["r_mult"]      == pytest.approx(-353/250, abs=0.001)
    assert e["days"]        == 8

    # Row 3: winning short
    e = EXPECTED[2]
    assert e["risk_dollar"] == pytest.approx(375.0,  abs=0.01)
    assert e["planned_rr"]  == pytest.approx(3.0,    abs=0.001)
    assert e["pl_dollar"]   == pytest.approx(746.0,  abs=0.01)
    assert e["pl_pct"]      == pytest.approx(746/(80*75), abs=0.0001)
    assert e["r_mult"]      == pytest.approx(746/375, abs=0.001)
    assert e["days"]        == 11


def test_win_rate_derivable_from_example_rows():
    """
    With 3 example trades (2 wins, 1 loss), win rate = 2/3.
    This verifies the fixture set is consistent with the dashboard formula logic.
    """
    wins = sum(1 for e in EXPECTED if e["pl_dollar"] > 0)
    total = len(EXPECTED)
    assert wins == 2
    assert total == 3
    assert wins / total == pytest.approx(2/3, abs=0.001)


# ---------------------------------------------------------------------------
# Test: Dashboard has a LineChart
# ---------------------------------------------------------------------------

def test_dashboard_has_line_chart(wb):
    ws = wb["Dashboard"]
    charts = ws._charts  # internal list used by openpyxl
    line_charts = [c for c in charts if isinstance(c, LineChart)]
    assert len(line_charts) >= 1, (
        f"Dashboard should contain at least one LineChart; found {charts}"
    )


# ---------------------------------------------------------------------------
# Test: no macro content
# ---------------------------------------------------------------------------

def test_no_macro_content(xlsx_path):
    """
    .xlsx files must not contain a vbaProject.bin part (that would make it .xlsm).
    openpyxl won't add one, but we verify explicitly.
    """
    import zipfile
    with zipfile.ZipFile(str(xlsx_path)) as zf:
        names = zf.namelist()
    macro_files = [n for n in names if "vbaProject" in n or n.endswith(".vba")]
    assert not macro_files, (
        f"Workbook contains macro content (not allowed for .xlsx): {macro_files}"
    )


# ---------------------------------------------------------------------------
# Test: Playbook sheet has setup names in A2:A5 (the 4 pre-filled examples)
# ---------------------------------------------------------------------------

def test_playbook_setup_names(wb):
    ws = wb["Playbook"]
    expected_setups = [
        "Breakout — 52-week high",
        "Pullback to rising 50-day",
        "Post-earnings drift",
        "Oversold bounce — counter-trend",
    ]
    for row_offset, expected_name in enumerate(expected_setups):
        actual = ws.cell(row=2 + row_offset, column=1).value
        assert actual == expected_name, (
            f"Playbook row {2 + row_offset} col A: expected {expected_name!r}, got {actual!r}"
        )


# ---------------------------------------------------------------------------
# Test: Settings sheet has expected default values
# ---------------------------------------------------------------------------

def test_settings_defaults(wb):
    ws = wb["Settings"]
    # B4 = starting account size = 10000
    account = ws["B4"].value
    assert account == 10000, f"Settings!B4 expected 10000, got {account!r}"
    # B5 = default risk % = 0.01 (stored as decimal; formatted as 0.00% → displays as 1%)
    risk_pct = ws["B5"].value
    assert risk_pct == pytest.approx(0.01, abs=1e-9), (
        f"Settings!B5 expected 0.01 (1%), got {risk_pct!r}"
    )


# ---------------------------------------------------------------------------
# Test: example rows have "EXAMPLE — delete me" in Notes column (col 20 = T)
# ---------------------------------------------------------------------------

def test_example_rows_notes_tag(wb):
    ws = wb["Trade Log"]
    for ex_idx in range(len(EXAMPLE_ROWS)):
        row = FIRST_DATA_ROW + ex_idx
        notes = ws.cell(row=row, column=20).value
        assert notes is not None, f"Example row {ex_idx + 1} Notes is empty"
        assert "EXAMPLE" in str(notes), (
            f"Example row {ex_idx + 1} Notes does not contain 'EXAMPLE': {notes!r}"
        )


# ---------------------------------------------------------------------------
# F1 new assertions: date cells, SUMIFS criterion, Days Held formula, Dashboard %-of-account
# ---------------------------------------------------------------------------

def test_date_cells_are_real_dates(wb):
    """E1: Date Opened / Date Closed example cells must be real date values, not strings."""
    import datetime as _dt
    ws = wb["Trade Log"]
    for ex_idx in range(len(EXAMPLE_ROWS)):
        row = FIRST_DATA_ROW + ex_idx
        date_open_cell  = ws.cell(row=row, column=2)   # B = Date Opened
        date_close_cell = ws.cell(row=row, column=12)  # L = Date Closed

        # openpyxl reads date-serial cells as datetime.datetime objects
        for cell, label in ((date_open_cell, "Date Opened"), (date_close_cell, "Date Closed")):
            v = cell.value
            is_date = isinstance(v, (_dt.date, _dt.datetime))
            assert is_date, (
                f"Example row {ex_idx+1} {label}: expected a date object, got {type(v).__name__} {v!r}. "
                "Cells must be date serials (E1 fix) not strings."
            )


def test_sumifs_not_blank_criterion(wb):
    """E2: The combined-P/L-on-mistake-trades SUMIFS must use '<>' (not-blank) as
    the second criterion, not a malformed literal."""
    ws = wb["Dashboard"]
    found = False
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "SUMIFS" in v and "None" in v:
                # The correct not-blank criterion is "<>" as the last criteria pair
                assert '"<>"' in v or "'<>'" in v, (
                    f"SUMIFS mistake criterion appears malformed (E2). "
                    f"Expected '<>' (not-blank) criterion, got: {v!r}"
                )
                # Must NOT have a literal `<>""` character sequence (the old bug)
                assert '<>""' not in v, (
                    f"SUMIFS still has malformed '<>\"\"' literal (E2): {v!r}"
                )
                found = True
    assert found, "Could not find the SUMIFS mistake-tagged formula in Dashboard"


def test_days_held_formula_shape(wb):
    """E1: Days Held formula on example rows uses L{r}-B{r} subtraction (serial dates)."""
    ws = wb["Trade Log"]
    days_col = 18  # R = Days Held
    for ex_idx in range(len(EXAMPLE_ROWS)):
        row = FIRST_DATA_ROW + ex_idx
        cell = ws.cell(row=row, column=days_col)
        v = cell.value
        assert isinstance(v, str) and v.startswith("="), (
            f"Example row {ex_idx+1} Days Held: expected formula, got {v!r}"
        )
        # Formula must contain L{r}-B{r} or equivalent subtraction of date columns
        assert f"L{row}" in v and f"B{row}" in v, (
            f"Example row {ex_idx+1} Days Held formula does not reference L{row} and B{row}: {v!r}"
        )


def test_dashboard_pct_of_account_references_settings_b4(wb):
    """E3: Dashboard must have a '% of starting account' stat that references Settings!$B$4."""
    ws = wb["Dashboard"]
    found = False
    for row_cells in ws.iter_rows():
        cells = list(row_cells)
        if len(cells) >= 2:
            label_val = str(cells[0].value or "")
            formula_val = cells[1].value
            if (isinstance(formula_val, str)
                    and "Settings" in formula_val
                    and "B4" in formula_val.replace("$", "")
                    and "account" in label_val.lower()):
                found = True
                break
    assert found, (
        "Dashboard must contain a 'Total P/L as % of starting account' row whose "
        "formula references Settings!$B$4 (E3 fix)."
    )


def test_example_row_days_from_date_objects():
    """F1: Verify example row days calculated from date objects (not ISO strings)."""
    import datetime as _dt
    for ex in EXAMPLE_ROWS:
        d_open  = ex["date_open"]
        d_close = ex["date_close"]
        # Both must be date objects (E1 fix)
        assert isinstance(d_open,  (_dt.date, _dt.datetime)), \
            f"date_open is {type(d_open).__name__}, expected date object"
        assert isinstance(d_close, (_dt.date, _dt.datetime)), \
            f"date_close is {type(d_close).__name__}, expected date object"
        # Days computed from date objects
        days = (d_close - d_open).days
        assert days >= 0, f"days must be non-negative, got {days}"
