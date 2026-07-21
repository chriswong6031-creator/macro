"""
build_trading_journal_xlsx.py
Deterministic generator for site/assets/downloads/mastermind-trading-journal.xlsx.

Run:  python3 scripts/build_trading_journal_xlsx.py
Same inputs => identical bytes (openpyxl, no embedded timestamps).
Allowed formula functions: IF IFERROR AND OR ABS SUMIF SUMIFS COUNTIF COUNTIFS
                            AVERAGEIF AVERAGEIFS MAX MIN SUM COUNT
No macros, no XLOOKUP, no LET, no LAMBDA, no defined names.
"""

from __future__ import annotations

import datetime
import io
import os
import re
import zipfile
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

OUTPUT_PATH = Path("site/assets/downloads/mastermind-trading-journal.xlsx")

# Fixed creation/modification date for byte-determinism.
FIXED_DATE = datetime.datetime(2026, 7, 20, 0, 0, 0)

# Colours
DARK_HEADER = "1F2D3D"      # dark navy for header bands
WHITE       = "FFFFFF"
ROW_TINT_A  = "F5F7FA"      # alternating row light tint
ROW_TINT_B  = "FFFFFF"
ACCENT_BLUE = "2E5BE8"
GREY_HELPER = "D9D9D9"      # light grey for helper columns

# Setups (must be ≤20 rows; used in Playbook and Trade Log dropdown)
SETUPS = [
    "Breakout — 52-week high",
    "Pullback to rising 50-day",
    "Post-earnings drift",
    "Oversold bounce — counter-trend",
    "",  # rows 5-20 intentionally blank for user entries
    "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
]
SETUPS = SETUPS[:20]  # ensure exactly 20

MISTAKE_OPTS = [
    "None",
    "Chased entry",
    "Moved stop",
    "Oversized",
    "Early exit",
    "No plan",
    "Revenge trade",
    "Other",
]

DIRECTION_OPTS = ["Long", "Short"]

# Trade Log column definitions
# Each entry: (header label, width, number_format)
TRADE_LOG_COLS = [
    ("#",               5,  "General"),
    ("Date Opened",    13,  "YYYY-MM-DD"),
    ("Ticker",          8,  "General"),
    ("Direction",      10,  "General"),
    ("Setup",          28,  "General"),
    ("Entry",          10,  "#,##0.00"),
    ("Stop",           10,  "#,##0.00"),
    ("Target",         10,  "#,##0.00"),
    ("Shares",          8,  "#,##0"),
    ("Risk $",         10,  "#,##0.00"),        # col 10 = J
    ("Planned R:R",    12,  "0.00"),             # col 11 = K
    ("Date Closed",    13,  "YYYY-MM-DD"),       # col 12 = L
    ("Exit",           10,  "#,##0.00"),         # col 13 = M
    ("Fees",           10,  "#,##0.00"),         # col 14 = N
    ("P/L $",          12,  "#,##0.00;[Red]-#,##0.00"),  # col 15 = O
    ("P/L %",          10,  "0.00%;[Red]-0.00%"),         # col 16 = P
    ("R Multiple",     12,  "0.00;[Red]-0.00"),            # col 17 = Q
    ("Days Held",       9,  "0"),                # col 18 = R
    ("Mistake?",       17,  "General"),          # col 19 = S
    ("Notes",          30,  "General"),          # col 20 = T
]

# Helper columns (far right, lightly greyed) — cumulative P/L for chart
HELPER_CUM_PL_COL  = 22  # V — cumulative P/L (used by chart)
HELPER_TRADE_N_COL = 21  # U — trade sequence number (used by chart x-axis)

# Column letter shortcuts (1-indexed)
COL_NUM       =  1   # A
COL_DATE_OPEN =  2   # B
COL_TICKER    =  3   # C
COL_DIR       =  4   # D
COL_SETUP     =  5   # E
COL_ENTRY     =  6   # F
COL_STOP      =  7   # G
COL_TARGET    =  8   # H
COL_SHARES    =  9   # I
COL_RISK      = 10   # J
COL_PLANRR    = 11   # K
COL_DATE_CLS  = 12   # L
COL_EXIT      = 13   # M
COL_FEES      = 14   # N
COL_PL_DOLLAR = 15   # O
COL_PL_PCT    = 16   # P
COL_R_MULT    = 17   # Q
COL_DAYS      = 18   # R
COL_MISTAKE   = 19   # S
COL_NOTES     = 20   # T

DATA_ROWS = 600   # formula-ready rows after header
HEADER_ROW = 1
FIRST_DATA = 2

# Example rows (three rows of realistic example trades)
# Values chosen so every formula column is exercised and produces
# sensible non-zero results.  Dates are real datetime.date objects (E1 fix)
# so openpyxl writes date serials; the Days Held formula (L{r}-B{r}) then
# works on serials in both Excel and Google Sheets.
#
# Row 1: winning long
# Row 2: losing long
# Row 3: winning short
#
# Exact fixture values used in tests:
#   Row 1: Entry=50, Stop=47, Target=59, Shares=100, Fees=5,
#           Exit=57, DateOpen=2026-06-01, DateClose=2026-06-15
#     Risk$ = (50-47)*100 = 300
#     PlannedRR = (59-50)/(50-47) = 3.00
#     PL$ = (57-50)*100 - 5 = 695
#     PL% = 695/(50*100) = 0.139 = 13.9%
#     R = 695/300 = 2.3167
#     Days = 14
#
#   Row 2: Entry=120, Stop=115, Target=135, Shares=50, Fees=3,
#           Exit=113, DateOpen=2026-06-10, DateClose=2026-06-18
#     Risk$ = (120-115)*50 = 250
#     PlannedRR = (135-120)/(120-115) = 3.00
#     PL$ = (113-120)*50 - 3 = -353
#     PL% = -353/(120*50) = -0.05883
#     R = -353/250 = -1.412
#     Days = 8
#
#   Row 3: winning short
#     Entry=80, Stop=85, Target=65, Shares=75, Fees=4,
#     Exit=70, DateOpen=2026-06-20, DateClose=2026-07-01
#     Risk$ = (85-80)*75 = 375
#     PlannedRR = (80-65)/(85-80) = 3.00
#     PL$ = (80-70)*75 - 4 = 746
#     PL% = 746/(80*75) = 0.12433
#     R = 746/375 = 1.9893
#     Days = 11

# E1: dates as real date objects (openpyxl writes date serials)
EXAMPLE_ROWS = [
    {
        "num": 1,
        "date_open": datetime.date(2026, 6, 1),
        "ticker": "AAPL",
        "direction": "Long",
        "setup": "Breakout — 52-week high",
        "entry": 50.00,
        "stop": 47.00,
        "target": 59.00,
        "shares": 100,
        "date_close": datetime.date(2026, 6, 15),
        "exit": 57.00,
        "fees": 5.00,
        "mistake": "None",
        "notes": "EXAMPLE — delete me",
    },
    {
        "num": 2,
        "date_open": datetime.date(2026, 6, 10),
        "ticker": "MSFT",
        "direction": "Long",
        "setup": "Pullback to rising 50-day",
        "entry": 120.00,
        "stop": 115.00,
        "target": 135.00,
        "shares": 50,
        "date_close": datetime.date(2026, 6, 18),
        "exit": 113.00,
        "fees": 3.00,
        "mistake": "Moved stop",
        "notes": "EXAMPLE — delete me",
    },
    {
        "num": 3,
        "date_open": datetime.date(2026, 6, 20),
        "ticker": "SPY",
        "direction": "Short",
        "setup": "Post-earnings drift",
        "entry": 80.00,
        "stop": 85.00,
        "target": 65.00,
        "shares": 75,
        "date_close": datetime.date(2026, 7, 1),
        "exit": 70.00,
        "fees": 4.00,
        "mistake": "None",
        "notes": "EXAMPLE — delete me",
    },
]


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _header_style(cell, text: str) -> None:
    """Apply dark header band style to a cell."""
    cell.value = text
    cell.fill = _fill(DARK_HEADER)
    cell.font = _font(bold=True, color=WHITE, size=11)
    cell.alignment = _center()
    cell.border = _thin_border()


def _subheader_style(cell, text: str) -> None:
    """Lighter sub-header (used on Dashboard sections)."""
    cell.value = text
    cell.fill = _fill("2E5BE8")
    cell.font = _font(bold=True, color=WHITE, size=10)
    cell.alignment = _left()
    cell.border = _thin_border()


def _label_style(cell, text: str) -> None:
    cell.value = text
    cell.font = _font(bold=True, size=10)
    cell.alignment = _left()
    cell.border = _thin_border()


def _data_style(cell, value=None, fmt: str = "General", row_idx: int = 0) -> None:
    """Apply alternating row tint + number format."""
    if value is not None:
        cell.value = value
    cell.number_format = fmt
    tint = ROW_TINT_A if row_idx % 2 == 0 else ROW_TINT_B
    cell.fill = _fill(tint)
    cell.alignment = _left()
    cell.border = _thin_border()


def _helper_style(cell, value=None, fmt: str = "General") -> None:
    """Light grey helper column cell."""
    if value is not None:
        cell.value = value
    cell.number_format = fmt
    cell.fill = _fill(GREY_HELPER)
    cell.font = _font(color="666666", size=9)
    cell.alignment = _center()
    cell.border = _thin_border()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_start_here(wb: Workbook) -> None:
    ws = wb.create_sheet("Start Here")

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    row = 1
    # Title
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws.cell(row=row, column=1)
    cell.value = "Mastermind Trading Journal — Quick Start Guide"
    cell.fill = _fill(DARK_HEADER)
    cell.font = _font(bold=True, color=WHITE, size=14)
    cell.alignment = _center()
    row += 2

    # What each sheet does
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    c.value = "WHAT EACH SHEET DOES"
    c.fill = _fill(ACCENT_BLUE)
    c.font = _font(bold=True, color=WHITE, size=11)
    c.alignment = _left()
    row += 1

    sheet_desc = [
        ("Start Here",  "This guide. Read once."),
        ("Trade Log",   "One row per trade. Log every trade here — closed or open."),
        ("Settings",    "Set your account size and default risk % here. Account size feeds the Dashboard '% of account' row; risk % feeds the Risk $ helper."),
        ("Dashboard",   "All stats auto-calculated from your Trade Log. No manual entry needed."),
        ("Playbook",    "Your personal setup library. Name and define up to 20 setups; "
                        "Trade Log dropdown pulls from this list."),
    ]
    for name, desc in sheet_desc:
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=1).font = _font(bold=True, size=10)
        ws.cell(row=row, column=2).value = desc
        ws.cell(row=row, column=2).font = _font(size=10)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    # How to log a trade in 30 seconds
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    c.value = "HOW TO LOG A TRADE IN 30 SECONDS"
    c.fill = _fill(ACCENT_BLUE)
    c.font = _font(bold=True, color=WHITE, size=11)
    c.alignment = _left()
    row += 1

    steps = [
        "1. Go to Trade Log.",
        "2. Fill in: Date Opened, Ticker, Direction (Long/Short), Setup (pick from dropdown).",
        "3. Enter Entry, Stop, Target, and Shares — Risk $ and Planned R:R calculate automatically.",
        "4. When the trade closes: fill in Date Closed, Exit, and Fees.",
        "   P/L $, P/L %, R Multiple, and Days Held all populate automatically.",
        "5. Tag any mistake from the Mistake? dropdown. That feeds the Dashboard 'cost of mistakes' line.",
    ]
    for step in steps:
        ws.cell(row=row, column=1).value = step
        ws.cell(row=row, column=1).font = _font(size=10)
        ws.merge_cells(f"A{row}:B{row}")
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    # What R multiples are
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    c.value = "WHAT R MULTIPLES ARE (AND WHY THE DASHBOARD USES THEM)"
    c.fill = _fill(ACCENT_BLUE)
    c.font = _font(bold=True, color=WHITE, size=11)
    c.alignment = _left()
    row += 1

    r_text = [
        ("R = 1.0",
         "means you made exactly what you risked on the trade (P/L $ equals your Risk $)."),
        ("R = 2.0",
         "means you made twice your risk. A trade where you risked $300 and made $600."),
        ("R = -1.0",
         "means you lost your full planned risk — a clean, disciplined stop."),
        ("Why R?",
         "Dollar P/L varies by position size. R normalises performance across different "
         "account sizes and position sizes, so you can compare setups fairly and calculate "
         "expectancy (average R per trade) across your whole journal."),
        ("Expectancy",
         "Win rate x avg winning R + Loss rate x avg losing R. Positive expectancy means "
         "your edge is real over a large sample of trades."),
    ]
    for label, desc in r_text:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = _font(bold=True, size=10)
        ws.cell(row=row, column=2).value = desc
        ws.cell(row=row, column=2).font = _font(size=10)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    # Example rows note
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    c.value = "THREE EXAMPLE ROWS IN THE TRADE LOG"
    c.fill = _fill(ACCENT_BLUE)
    c.font = _font(bold=True, color=WHITE, size=11)
    c.alignment = _left()
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    c.value = (
        "Rows 2–4 of the Trade Log are pre-filled with example trades (a winning long, "
        "a losing long, and a winning short). They show every formula working with real "
        "numbers. Delete them before you start logging your own trades — they will skew "
        "your Dashboard stats otherwise. Notes column says 'EXAMPLE — delete me'."
    )
    c.font = _font(size=10)
    c.alignment = Alignment(wrap_text=True)
    row += 2

    # Disclaimer
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Educational tool — not investment advice."
    c.fill = _fill("FFF3CD")
    c.font = _font(bold=True, color="856404", size=10)
    c.alignment = _center()

    ws.sheet_view.showGridLines = False


def _build_settings(wb: Workbook) -> None:
    ws = wb.create_sheet("Settings")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40

    row = 1
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Settings — Account & Risk Defaults"
    c.fill = _fill(DARK_HEADER)
    c.font = _font(bold=True, color=WHITE, size=13)
    c.alignment = _center()
    row += 2

    ws.cell(row=row, column=1).value = "Label"
    ws.cell(row=row, column=2).value = "Value"
    ws.cell(row=row, column=3).value = "Notes"
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = _fill(DARK_HEADER)
        ws.cell(row=row, column=col).font = _font(bold=True, color=WHITE, size=10)
        ws.cell(row=row, column=col).alignment = _center()
        ws.cell(row=row, column=col).border = _thin_border()
    row += 1

    settings_data = [
        ("Starting account size",  10000,  "#,##0.00", "Change this to your actual account size in your currency."),
        ("Default risk % per trade", 0.01, "0.00%",    "1% means you risk 1% of account per trade at your full stop."),
        ("Currency label",          "USD",  "General",  "Display label only. Does not convert values."),
    ]

    for label, val, fmt, note in settings_data:
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = _font(bold=True, size=10)
        ws.cell(row=row, column=1).border = _thin_border()
        ws.cell(row=row, column=1).alignment = _left()

        c = ws.cell(row=row, column=2)
        c.value = val
        c.number_format = fmt
        c.font = _font(bold=True, color=ACCENT_BLUE, size=11)
        c.border = _thin_border()
        c.alignment = _center()

        ws.cell(row=row, column=3).value = note
        ws.cell(row=row, column=3).font = _font(size=10)
        ws.cell(row=row, column=3).border = _thin_border()
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        row += 1

    row += 1
    # Helper: risk $ at default
    ws.cell(row=row, column=1).value = "Risk $ at default (account × risk %)"
    ws.cell(row=row, column=1).font = _font(bold=True, size=10)
    ws.cell(row=row, column=1).border = _thin_border()
    ws.cell(row=row, column=1).alignment = _left()

    c = ws.cell(row=row, column=2)
    c.value = "=B4*B5"
    c.number_format = "#,##0.00"
    c.font = _font(bold=True, color=ACCENT_BLUE, size=11)
    c.border = _thin_border()
    c.alignment = _center()

    ws.cell(row=row, column=3).value = "This is the dollar amount you would risk on a max-size trade at your defaults."
    ws.cell(row=row, column=3).font = _font(size=10)
    ws.cell(row=row, column=3).border = _thin_border()
    ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    row += 2

    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1)
    c.value = (
        "Settings!$B$4 (account size) feeds the Dashboard 'Total P/L as % of starting account' row "
        "and the Risk $ helper below. Settings!$B$5 (risk %) feeds the Risk $ helper. "
        "You can change these values at any time — all dependent cells update automatically."
    )
    c.font = _font(size=10, color="666666")
    c.alignment = Alignment(wrap_text=True)

    ws.sheet_view.showGridLines = False


def _trade_log_formulas(row: int) -> dict[int, str]:
    """Return formula strings for computed columns at a given data row."""
    r = row
    # Risk $ = |Entry - Stop| * Shares, blank until Entry and Stop present
    risk = (
        f"=IFERROR(IF(AND(F{r}<>\"\",G{r}<>\"\",I{r}<>\"\"),"
        f"ABS(F{r}-G{r})*I{r},\"\"),\"\")"
    )
    # Planned R:R = |Target - Entry| / |Entry - Stop|
    planrr = (
        f"=IFERROR(IF(AND(F{r}<>\"\",G{r}<>\"\",H{r}<>\"\"),"
        f"ABS(H{r}-F{r})/ABS(F{r}-G{r}),\"\"),\"\")"
    )
    # P/L $: Long = (Exit-Entry)*Shares-Fees, Short = (Entry-Exit)*Shares-Fees
    # Blank until Exit present
    pl_dollar = (
        f"=IFERROR(IF(AND(M{r}<>\"\",F{r}<>\"\",I{r}<>\"\"),"
        f"IF(D{r}=\"Long\",(M{r}-F{r})*I{r}-IF(N{r}<>\"\",N{r},0),"
        f"IF(D{r}=\"Short\",(F{r}-M{r})*I{r}-IF(N{r}<>\"\",N{r},0),\"\")),"
        f"\"\"),\"\")"
    )
    # P/L % = P/L $ / (Entry * Shares)
    pl_pct = (
        f"=IFERROR(IF(AND(O{r}<>\"\",F{r}<>\"\",I{r}<>\"\"),"
        f"O{r}/(F{r}*I{r}),\"\"),\"\")"
    )
    # R Multiple = P/L $ / Risk $
    r_mult = (
        f"=IFERROR(IF(AND(O{r}<>\"\",J{r}<>\"\",J{r}<>0),"
        f"O{r}/J{r},\"\"),\"\")"
    )
    # Days Held = Date Closed - Date Opened
    days = (
        f"=IFERROR(IF(AND(L{r}<>\"\",B{r}<>\"\"),"
        f"L{r}-B{r},\"\"),\"\")"
    )
    return {
        COL_RISK:      risk,
        COL_PLANRR:    planrr,
        COL_PL_DOLLAR: pl_dollar,
        COL_PL_PCT:    pl_pct,
        COL_R_MULT:    r_mult,
        COL_DAYS:      days,
    }


def _build_trade_log(wb: Workbook) -> None:
    ws = wb.create_sheet("Trade Log")

    # --- Column widths and number formats ---
    for col_idx, (header, width, fmt) in enumerate(TRADE_LOG_COLS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

    # Helper columns
    ws.column_dimensions[get_column_letter(HELPER_TRADE_N_COL)].width = 10
    ws.column_dimensions[get_column_letter(HELPER_CUM_PL_COL)].width = 14

    # --- Header row ---
    for col_idx, (header, width, fmt) in enumerate(TRADE_LOG_COLS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        _header_style(cell, header)

    # Helper column headers
    h_n = ws.cell(row=HEADER_ROW, column=HELPER_TRADE_N_COL)
    _helper_style(h_n, "Trade #")
    h_cum = ws.cell(row=HEADER_ROW, column=HELPER_CUM_PL_COL)
    _helper_style(h_cum, "Cumulative P/L")

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Dropdowns (data validation) ---
    # Direction dropdown: D2:D601
    dv_dir = DataValidation(
        type="list",
        formula1='"Long,Short"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_dir.sqref = f"D{FIRST_DATA}:D{FIRST_DATA + DATA_ROWS - 1}"
    dv_dir.error = "Choose Long or Short."
    dv_dir.errorTitle = "Invalid direction"
    ws.add_data_validation(dv_dir)

    # Setup dropdown: E2:E601 — pulls from Playbook A2:A21
    dv_setup = DataValidation(
        type="list",
        formula1="Playbook!$A$2:$A$21",
        allow_blank=True,
        showDropDown=False,
    )
    dv_setup.sqref = f"E{FIRST_DATA}:E{FIRST_DATA + DATA_ROWS - 1}"
    ws.add_data_validation(dv_setup)

    # Mistake dropdown: S2:S601
    mistake_joined = ",".join(MISTAKE_OPTS)
    dv_mistake = DataValidation(
        type="list",
        formula1=f'"{mistake_joined}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_mistake.sqref = f"S{FIRST_DATA}:S{FIRST_DATA + DATA_ROWS - 1}"
    ws.add_data_validation(dv_mistake)

    # --- Example rows ---
    for ex_idx, ex in enumerate(EXAMPLE_ROWS):
        row = FIRST_DATA + ex_idx
        tint = ROW_TINT_A if ex_idx % 2 == 0 else ROW_TINT_B

        formulas = _trade_log_formulas(row)

        static_values = {
            COL_NUM:       ex["num"],
            COL_DATE_OPEN: ex["date_open"],
            COL_TICKER:    ex["ticker"],
            COL_DIR:       ex["direction"],
            COL_SETUP:     ex["setup"],
            COL_ENTRY:     ex["entry"],
            COL_STOP:      ex["stop"],
            COL_TARGET:    ex["target"],
            COL_SHARES:    ex["shares"],
            COL_DATE_CLS:  ex["date_close"],
            COL_EXIT:      ex["exit"],
            COL_FEES:      ex["fees"],
            COL_MISTAKE:   ex["mistake"],
            COL_NOTES:     ex["notes"],
        }

        for col_idx, (header, width, fmt) in enumerate(TRADE_LOG_COLS, start=1):
            cell = ws.cell(row=row, column=col_idx)
            if col_idx in formulas:
                cell.value = formulas[col_idx]
            elif col_idx in static_values:
                cell.value = static_values[col_idx]
            cell.number_format = fmt
            cell.fill = _fill(tint)
            cell.alignment = _left()
            cell.border = _thin_border()

        # Helper columns for example rows
        _helper_style(
            ws.cell(row=row, column=HELPER_TRADE_N_COL),
            value=ex_idx + 1,
            fmt="0",
        )
        # Cumulative P/L formula
        cum_cell = ws.cell(row=row, column=HELPER_CUM_PL_COL)
        o_col = get_column_letter(COL_PL_DOLLAR)
        if ex_idx == 0:
            cum_cell.value = f"=IF({o_col}{row}<>\"\",{o_col}{row},0)"
        else:
            prev_row = row - 1
            prev_cum = get_column_letter(HELPER_CUM_PL_COL)
            cum_cell.value = (
                f"=IF({o_col}{row}<>\"\",{prev_cum}{prev_row}+{o_col}{row},{prev_cum}{prev_row})"
            )
        _helper_style(cum_cell, fmt="#,##0.00")

    # --- Blank formula rows (rows 5 to 601) ---
    for data_row_idx in range(len(EXAMPLE_ROWS), DATA_ROWS):
        row = FIRST_DATA + data_row_idx
        tint = ROW_TINT_A if data_row_idx % 2 == 0 else ROW_TINT_B
        formulas = _trade_log_formulas(row)

        for col_idx, (header, width, fmt) in enumerate(TRADE_LOG_COLS, start=1):
            cell = ws.cell(row=row, column=col_idx)
            if col_idx in formulas:
                cell.value = formulas[col_idx]
            cell.number_format = fmt
            cell.fill = _fill(tint)
            cell.alignment = _left()
            cell.border = _thin_border()

        # Helper cols
        h_n_cell = ws.cell(row=row, column=HELPER_TRADE_N_COL)
        h_n_cell.value = (
            f"=IF({get_column_letter(COL_PL_DOLLAR)}{row}<>\"\",{data_row_idx + 1},\"\")"
        )
        _helper_style(h_n_cell, fmt="0")

        cum_cell = ws.cell(row=row, column=HELPER_CUM_PL_COL)
        o_col = get_column_letter(COL_PL_DOLLAR)
        prev_row = row - 1
        prev_cum = get_column_letter(HELPER_CUM_PL_COL)
        cum_cell.value = (
            f"=IF({o_col}{row}<>\"\",{prev_cum}{prev_row}+{o_col}{row},{prev_cum}{prev_row})"
        )
        _helper_style(cum_cell, fmt="#,##0.00")

    ws.sheet_view.showGridLines = False


def _build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12

    ws.freeze_panes = "A1"

    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1)
    c.value = "Trading Dashboard — auto-calculated from your Trade Log"
    c.fill = _fill(DARK_HEADER)
    c.font = _font(bold=True, color=WHITE, size=13)
    c.alignment = _center()
    row += 1

    # Helper: closed-trade range definition
    # "Closed" means P/L $ column (O) is not blank
    # Trade Log data: O2:O601
    o_range = f"'Trade Log'!$O$2:$O${FIRST_DATA + DATA_ROWS - 1}"
    q_range = f"'Trade Log'!$Q$2:$Q${FIRST_DATA + DATA_ROWS - 1}"  # R Multiple
    e_range = f"'Trade Log'!$E$2:$E${FIRST_DATA + DATA_ROWS - 1}"  # Setup
    s_range = f"'Trade Log'!$S$2:$S${FIRST_DATA + DATA_ROWS - 1}"  # Mistake
    r_range = f"'Trade Log'!$R$2:$R${FIRST_DATA + DATA_ROWS - 1}"  # Days Held

    def stat_row(label: str, formula: str, fmt: str = "#,##0.00", note: str = "") -> None:
        nonlocal row
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = _font(size=10)
        ws.cell(row=row, column=1).border = _thin_border()
        ws.cell(row=row, column=1).alignment = _left()

        c = ws.cell(row=row, column=2)
        c.value = formula
        c.number_format = fmt
        c.font = _font(bold=True, size=11, color="1F2D3D")
        c.border = _thin_border()
        c.alignment = _center()

        if note:
            ws.cell(row=row, column=3).value = note
            ws.cell(row=row, column=3).font = _font(size=9, color="666666")
            ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        row += 1

    def section_header(label: str) -> None:
        nonlocal row
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1)
        c.value = label
        c.fill = _fill(ACCENT_BLUE)
        c.font = _font(bold=True, color=WHITE, size=10)
        c.alignment = _left()
        row += 1

    # --- Section 1: Trade counts ---
    section_header("TRADE SUMMARY")
    stat_row("Total closed trades",
             f"=COUNTIF({o_range},\"<>\"&\"\")",
             "0",
             "Closed = rows where P/L $ is present")
    stat_row("Wins (P/L $ > 0)",
             f"=COUNTIF({o_range},\">0\")",
             "0")
    stat_row("Losses (P/L $ < 0)",
             f"=COUNTIF({o_range},\"<0\")",
             "0")
    stat_row("Win rate",
             f"=IFERROR(COUNTIF({o_range},\">0\")/COUNTIF({o_range},\"<>\"&\"\"),0)",
             "0.00%")
    row += 1

    # --- Section 2: P/L ---
    section_header("PROFIT & LOSS")
    stat_row("Total P/L $",
             f"=IFERROR(SUM({o_range}),0)",
             "#,##0.00;[Red]-#,##0.00")
    stat_row("Average win $",
             f"=IFERROR(AVERAGEIF({o_range},\">0\"),0)",
             "#,##0.00")
    stat_row("Average loss $",
             f"=IFERROR(AVERAGEIF({o_range},\"<0\"),0)",
             "#,##0.00;[Red]-#,##0.00")
    stat_row("Best trade $",
             f"=IFERROR(MAX({o_range}),0)",
             "#,##0.00")
    stat_row("Worst trade $",
             f"=IFERROR(MIN({o_range}),0)",
             "#,##0.00;[Red]-#,##0.00")
    row += 1

    # --- Section 3: Edge stats ---
    section_header("EDGE STATISTICS")
    stat_row("Payoff ratio (avg win / |avg loss|)",
             f"=IFERROR(ABS(AVERAGEIF({o_range},\">0\")/AVERAGEIF({o_range},\"<0\")),0)",
             "0.00",
             "Above 1.0 means wins are larger than losses on average")
    stat_row("Profit factor (gross wins / |gross losses|)",
             f"=IFERROR(ABS(SUMIF({o_range},\">0\")/SUMIF({o_range},\"<0\")),0)",
             "0.00",
             "Above 1.0 means the system is profitable overall")
    stat_row("Expectancy per trade $",
             f"=IFERROR(SUM({o_range})/COUNTIF({o_range},\"<>\"&\"\"),0)",
             "#,##0.00",
             "Average dollar earned per closed trade")
    stat_row("Expectancy per trade (R)",
             f"=IFERROR(SUM({q_range})/COUNTIF({o_range},\"<>\"&\"\"),0)",
             "0.00",
             "Average R earned per closed trade")
    stat_row("Average R multiple",
             f"=IFERROR(AVERAGEIF({o_range},\"<>\"&\"\",{q_range}),0)",
             "0.00")
    stat_row("Average days held",
             f"=IFERROR(AVERAGEIF({o_range},\"<>\"&\"\",{r_range}),0)",
             "0.0")
    row += 1

    # --- Section 4: Cost of mistakes ---
    section_header("COST OF MISTAKES  (the feature TrendSpider's free journal doesn't have)")
    # "Mistake?" column S; count where mistake != "None" AND != ""
    # We count rows with a closed trade AND mistake is not "None"
    stat_row("Trades tagged as a mistake",
             f"=IFERROR(COUNTIFS({o_range},\"<>\"&\"\",{s_range},\"<>None\")-COUNTIFS({o_range},\"<>\"&\"\",{s_range},\"\"),0)",
             "0",
             "Closed trades where Mistake? is not None")
    stat_row("Combined P/L on mistake trades $",
             f"=IFERROR(SUMIFS({o_range},{s_range},\"<>None\",{s_range},\"<>\"),0)",
             "#,##0.00;[Red]-#,##0.00",
             "How much those tagged mistakes cost (or made) you")
    row += 1

    # --- Section 4b: Account context (references Settings!$B$4) ---
    section_header("ACCOUNT CONTEXT")
    # E3: Total P/L as % of starting account — references Settings!$B$4
    # IFERROR-guarded in case Settings!$B$4 is blank or zero
    stat_row("Total P/L as % of starting account",
             "=IFERROR(SUM(" + o_range + ")/Settings!$B$4,0)",
             "0.00%;[Red]-0.00%",
             "Total realised P/L relative to the account size in Settings!B4")
    row += 1

    # --- Section 5: Per-setup breakdown ---
    section_header("PER-SETUP BREAKDOWN")

    # Column headers for the breakdown table
    breakdown_headers = ["Setup", "Trades", "Win Rate", "Total P/L $", "Avg R"]
    for col_idx, h in enumerate(breakdown_headers, start=1):
        c = ws.cell(row=row, column=col_idx)
        c.value = h
        c.fill = _fill(DARK_HEADER)
        c.font = _font(bold=True, color=WHITE, size=10)
        c.alignment = _center()
        c.border = _thin_border()
    row += 1

    for setup_idx in range(1, 21):  # Playbook rows A2:A21
        setup_ref = f"Playbook!$A${setup_idx + 1}"

        # Only render if setup name is not blank
        setup_formula = f"=IFERROR(IF({setup_ref}=\"\",\"\",{setup_ref}),\"\")"
        trades_formula = (
            f"=IFERROR(IF({setup_ref}=\"\",\"\","
            f"COUNTIFS({e_range},{setup_ref},{o_range},\"<>\"&\"\")),\"\")"
        )
        winrate_formula = (
            f"=IFERROR(IF({setup_ref}=\"\",\"\","
            f"COUNTIFS({e_range},{setup_ref},{o_range},\">0\")"
            f"/COUNTIFS({e_range},{setup_ref},{o_range},\"<>\"&\"\")),\"\")"
        )
        total_pl_formula = (
            f"=IFERROR(IF({setup_ref}=\"\",\"\","
            f"SUMIFS({o_range},{e_range},{setup_ref})),\"\")"
        )
        avg_r_formula = (
            f"=IFERROR(IF({setup_ref}=\"\",\"\","
            f"AVERAGEIFS({q_range},{e_range},{setup_ref},{o_range},\"<>\"&\"\")),\"\")"
        )

        cells_data = [
            (1, setup_formula,    "General",                        _left()),
            (2, trades_formula,   "0",                              _center()),
            (3, winrate_formula,  "0.00%",                          _center()),
            (4, total_pl_formula, "#,##0.00;[Red]-#,##0.00",       _center()),
            (5, avg_r_formula,    "0.00;[Red]-0.00",               _center()),
        ]
        tint = ROW_TINT_A if setup_idx % 2 == 0 else ROW_TINT_B
        for col_idx, formula, fmt, align in cells_data:
            c = ws.cell(row=row, column=col_idx)
            c.value = formula
            c.number_format = fmt
            c.fill = _fill(tint)
            c.font = _font(size=10)
            c.alignment = align
            c.border = _thin_border()
        row += 1

    row += 1

    # --- Line chart: cumulative P/L ---
    section_header("CUMULATIVE P/L CHART")

    # Chart data lives in Trade Log helper columns U (trade#) and V (cumulative P/L)
    # We reference Trade Log!V2:V601 for the series values
    # and Trade Log!U2:U601 for x-axis categories
    chart_data_col = get_column_letter(HELPER_CUM_PL_COL)   # V
    chart_ref_start = FIRST_DATA
    chart_ref_end   = FIRST_DATA + DATA_ROWS - 1

    chart = LineChart()
    chart.title = "Cumulative P/L"
    chart.style = 10
    chart.y_axis.title = "Cumulative P/L ($)"
    chart.x_axis.title = "Trade #"
    chart.height = 14
    chart.width  = 28
    chart.grouping = "standard"
    chart.smooth = True

    # Data series: Trade Log col V (cumulative P/L)
    data_ref = Reference(
        wb["Trade Log"],
        min_col=HELPER_CUM_PL_COL,
        min_row=HEADER_ROW,      # includes header for series title
        max_row=chart_ref_end,
    )
    chart.add_data(data_ref, titles_from_data=True)

    # x-axis categories: Trade Log col U (trade sequence number)
    cats = Reference(
        wb["Trade Log"],
        min_col=HELPER_TRADE_N_COL,
        min_row=FIRST_DATA,
        max_row=chart_ref_end,
    )
    chart.set_categories(cats)

    # Style the line
    if chart.series:
        s = chart.series[0]
        s.graphicalProperties.line.solidFill = ACCENT_BLUE
        s.graphicalProperties.line.width = 20000  # 2pt in EMUs

    ws.add_chart(chart, f"A{row}")

    ws.sheet_view.showGridLines = False


def _build_playbook(wb: Workbook) -> None:
    ws = wb.create_sheet("Playbook")

    col_widths = [30, 40, 35, 30, 30]
    col_headers = ["Setup", "Entry Rules", "Exit Rules", "Valid Market Condition", "Notes"]

    for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Title
    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1)
    c.value = "Playbook — Your Personal Setup Library"
    c.fill = _fill(DARK_HEADER)
    c.font = _font(bold=True, color=WHITE, size=13)
    c.alignment = _center()

    # Header row (row 2 = Playbook!A2 referenced by Trade Log dropdown)
    for col_idx, header in enumerate(col_headers, start=1):
        c = ws.cell(row=2, column=col_idx)
        _header_style(c, header)

    ws.freeze_panes = "A3"

    # Pre-fill 4 example setups (rows 3-6 = A3:A6; dropdown ref is A2:A21)
    # Note: data goes in rows 2..21 where row 2 = header.
    # The dropdown formula is Playbook!$A$2:$A$21 which includes the header row.
    # To avoid the header "Setup" appearing as a selectable option, we put
    # setup names in rows 2-21, with row 2 being the column header *and* the
    # first setup slot only if we shift.  Per spec: "A2:A21" feeds the dropdown.
    # We put header in row 1 (title) and row 2 = first setup name slot.
    # Rebuild: title row 1, header row 2, data rows 3-22.
    # BUT the CONTRACT spec says "Playbook A2:A21" feeds dropdown.
    # So setup names must be in A2:A21. We'll put column labels in row 1 only.

    example_setups = [
        (
            "Breakout — 52-week high",
            "Price closes at or above 52-week high on above-average volume (>1.5x 20-day avg). "
            "Entry on next-day open or intraday retest of the breakout level.",
            "Scale out 50% at 1R above entry. Trail stop to break-even after first scale. "
            "Full exit on close below 10-day EMA.",
            "Trending market (S&P 500 above 50-day and 200-day MA). Avoid in choppy, "
            "range-bound markets or broad downtrend.",
            "Best in early-to-mid uptrend. Volume confirmation is non-negotiable.",
        ),
        (
            "Pullback to rising 50-day",
            "Stock in uptrend (price above 50-day and 200-day MA). Price pulls back "
            "to within 2% of 50-day MA. Wait for one green close that holds the 50-day. "
            "Entry on next open or limit at the 50-day.",
            "First target = prior swing high. Partial exit there. "
            "Hard stop = daily close below 50-day MA.",
            "Works best when broader market is in a confirmed uptrend and sector is strong. "
            "Avoid during earnings season unless entry is >2 weeks from earnings date.",
            "Lower-risk entry point than chasing breakouts. "
            "Quality names in leading sectors have highest reliability.",
        ),
        (
            "Post-earnings drift",
            "Stock gaps up or down >5% on earnings with heavy volume. "
            "Wait for the first full trading session to pass (let initial volatility settle). "
            "Enter in the direction of the gap on a pullback to the gap-day VWAP or "
            "the prior-close level.",
            "Target = measured move of gap size projected from entry. "
            "Stop = opposite side of gap-day range. Exit in full if price reverses "
            "to fill the gap.",
            "Most reliable when the earnings surprise was material (beat/miss >10%) "
            "and sector sentiment is aligned. Avoid when index is in a volatile regime.",
            "Drift can last 3-10 trading days. Size conservatively — "
            "post-earnings vol remains elevated.",
        ),
        (
            "Oversold bounce — counter-trend",
            "RSI(14) below 30 on daily chart. Price at or near a major support level "
            "(prior swing low, round number, 200-day MA). Wait for a bullish reversal candle "
            "(hammer, engulfing) to form. Entry on next open.",
            "Target = prior resistance or 50% retracement of recent decline. "
            "Tight stop just below the reversal candle low. Exit at target or on any close "
            "below entry.",
            "Works against the short-term trend. Only take in broadly oversold market "
            "conditions. Do not hold through a trend-continuation break.",
            "Smaller size than trend trades (counter-trend). "
            "This is a mean-reversion play, not a new-trend entry.",
        ),
    ]

    for row_offset, (setup_name, entry, exit_, condition, notes) in enumerate(example_setups):
        row = 2 + row_offset   # rows 2-5
        tint = ROW_TINT_A if row_offset % 2 == 0 else ROW_TINT_B
        values = [setup_name, entry, exit_, condition, notes]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx)
            c.value = val
            c.font = _font(size=10)
            c.fill = _fill(tint)
            c.border = _thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60

    # Blank rows 6-21 (A6:A21 are empty setup slots for user's own setups)
    for row_offset in range(4, 20):
        row = 2 + row_offset
        tint = ROW_TINT_A if row_offset % 2 == 0 else ROW_TINT_B
        for col_idx in range(1, 6):
            c = ws.cell(row=row, column=col_idx)
            c.fill = _fill(tint)
            c.border = _thin_border()
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

_MODIFIED_RE = re.compile(
    rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
    re.DOTALL,
)

# Canonical modified timestamp for core.xml (pinned for determinism)
_CORE_MODIFIED_FIXED = (
    b'<dcterms:modified xmlns:dcterms="http://purl.org/dc/terms/"'
    b' xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"'
    b' xsi:type="dcterms:W3CDTF">2026-07-20T00:00:00Z</dcterms:modified>'
)


def _rezip_deterministic(src_bytes: bytes, fixed_date: tuple) -> bytes:
    """
    Re-write an xlsx zip archive with:
    - All zip entry timestamps set to fixed_date.
    - The dcterms:modified in docProps/core.xml replaced with a pinned value.

    openpyxl's save_workbook() explicitly sets workbook.properties.modified =
    datetime.datetime.now(...) before writing, so the only reliable fix is to
    rewrite the XML content after the fact.
    """
    src_buf = io.BytesIO(src_bytes)
    dst_buf = io.BytesIO()
    with zipfile.ZipFile(src_buf, "r") as src_zf, \
         zipfile.ZipFile(dst_buf, "w", compression=zipfile.ZIP_DEFLATED) as dst_zf:
        for info in src_zf.infolist():
            data = src_zf.read(info.filename)
            if info.filename == "docProps/core.xml":
                data = _MODIFIED_RE.sub(_CORE_MODIFIED_FIXED, data)
            info.date_time = fixed_date
            dst_zf.writestr(info, data)
    return dst_buf.getvalue()


def build(output_path: Path = OUTPUT_PATH) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Pin workbook properties for byte-determinism
    wb.properties.created  = FIXED_DATE
    wb.properties.modified = FIXED_DATE
    wb.properties.creator  = "Mastermind Research"
    wb.properties.lastModifiedBy = "Mastermind Research"

    # Build sheets in order
    _build_start_here(wb)
    _build_trade_log(wb)
    _build_settings(wb)
    _build_dashboard(wb)
    _build_playbook(wb)

    # Save to memory first, then re-zip with fixed timestamps for determinism.
    buf = io.BytesIO()
    wb.save(buf)
    raw_bytes = buf.getvalue()

    fixed_zip_date = (
        FIXED_DATE.year, FIXED_DATE.month, FIXED_DATE.day,
        FIXED_DATE.hour, FIXED_DATE.minute, FIXED_DATE.second,
    )
    deterministic_bytes = _rezip_deterministic(raw_bytes, fixed_zip_date)

    output_path.write_bytes(deterministic_bytes)
    return output_path


if __name__ == "__main__":
    import sys

    path = Path(sys.argv[1]) if len(sys.argv) > 1 else OUTPUT_PATH
    result = build(path)
    size = result.stat().st_size
    print(f"Written: {result}  ({size:,} bytes)")
