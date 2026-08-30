"""Deterministic finite-snapshot admission for the licensed BPC JV corpus.

This module has no live-source capability.  It accepts only explicitly supplied
finite files, keeps source bytes separate from normalized product records, and
never emits private URLs or export-time market overlays into the public record.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO, StringIO
import csv
import json
from pathlib import Path
import re
from typing import Any, Callable, Mapping, Sequence
import unicodedata

from openpyxl import load_workbook
from lib.dataos.identity import IssuerMaster, VendorAliasTable


class SnapshotError(RuntimeError):
    """Bounded finite-snapshot failure; messages never include source rows."""

    def __init__(self, code: str) -> None:
        self.code = code
        super().__init__(code)


@dataclass(frozen=True, slots=True)
class InputSpec:
    input_id: str
    safe_name: str
    sha256: str
    byte_count: int
    media_type: str
    role: str


@dataclass(frozen=True, slots=True)
class AdmittedInput:
    spec: InputSpec
    path: Path
    data: bytes


@dataclass(frozen=True, slots=True)
class HistoricalFdaRepair:
    rows: tuple[dict[str, str], ...]
    row_count: int
    repaired_count: int
    raw_sha256: str


@dataclass(frozen=True, slots=True)
class NormalizedCorpus:
    events: tuple[dict[str, Any], ...]
    jsonl: bytes
    coverage: dict[str, Any]
    historical_fda_repair: HistoricalFdaRepair


class SnapshotIdentityResolver:
    """Resolve BPC ticker evidence through the existing Data OS readers.

    Cross-vendor agreement is required.  A bounded alias row can prove a
    time-scoped resolution; open-bounded catalog rows can only prove a current
    catalog relationship and are labeled accordingly.  IssuerMaster is
    current-only by contract, so its answer is never backdated.
    """

    def __init__(self, aliases: VendorAliasTable, issuers: IssuerMaster) -> None:
        self.aliases = aliases
        self.issuers = issuers

    def __call__(self, ticker: str | None, on: date) -> dict[str, object]:
        if not ticker:
            return _default_identity(ticker, on)
        matches = [
            row
            for row in self.aliases.rows
            if row.vendor_symbol == ticker and row.covers(on)
        ]
        security_ids = sorted({row.security_id for row in matches})
        if not security_ids:
            return _default_identity(ticker, on)
        if len(security_ids) != 1:
            return {
                "resolution_state": "ambiguous",
                "security_id": None,
                "issuer_id": None,
                "resolution_basis": "ambiguous",
                "issuer_relationship_state": "unavailable",
            }
        security_id = security_ids[0]
        has_bounded_evidence = any(
            row.valid_from is not None or row.valid_to is not None for row in matches
        )
        issuer_id = self.issuers.issuer_of_security(security_id)
        return {
            "resolution_state": "resolved",
            "security_id": security_id,
            "issuer_id": issuer_id,
            "resolution_basis": "time_scoped_alias" if has_bounded_evidence else "current_catalog_only",
            "issuer_relationship_state": "current_only" if issuer_id else "unavailable",
        }


def identity_resolver_from_parquet_bytes(
    security_master: bytes,
    vendor_aliases: bytes,
) -> SnapshotIdentityResolver:
    """Load the two existing Data OS artifacts without minting another store."""

    try:
        import pandas as pd  # noqa: PLC0415

        master_records = pd.read_parquet(BytesIO(security_master)).to_dict("records")
        alias_records = pd.read_parquet(BytesIO(vendor_aliases)).to_dict("records")
        return SnapshotIdentityResolver(
            VendorAliasTable.from_records(alias_records),
            IssuerMaster.from_records(master_records),
        )
    except Exception:
        raise SnapshotError("IDENTITY_ARTIFACT_INVALID") from None


AUTHORIZED_INPUTS: dict[str, InputSpec] = {
    "workbook_w4": InputSpec(
        "workbook_w4",
        "BioPharmCatalyst_Tables.xlsx",
        "946c5f725ebfd3b71d254f229e006ba055a868a1d5d02d3344a74efb3882b535",
        353_040,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "canonical_workbook",
    ),
    "all_companies": InputSpec(
        "all_companies",
        "BioPharmCatalyst_All_Companies_Sorted_By_Ticker.csv",
        "a08afff0430c06138997f6b8a3e28fee63bb742eecdb4ea936c8bea99f225ee0",
        86_364,
        "text/csv",
        "identity_reconciliation",
    ),
    "historical_fda": InputSpec(
        "historical_fda",
        "biopharmcatalyst_historical_fda_all_verified_2009_2026.csv",
        "f3852d34aad9b65d95e31db807f9509cfb84770eb91998533cb3687cea3d9002",
        5_630_777,
        "text/csv",
        "historical_event_source",
    ),
    "mergers_acquisitions": InputSpec(
        "mergers_acquisitions",
        "biopharmcatalyst_mergers_acquisitions.csv",
        "aa33b6dea553b982b32621a3ee759d20283c25b1e6d267289f6e7d38e5afb3fd",
        268_490,
        "text/csv",
        "inventory_only",
    ),
    "hedge_funds": InputSpec(
        "hedge_funds",
        "biopharmcatalyst_hedge_funds.csv",
        "fbb968bae5f4f5f6a33f21ee6c02db4450f26cf19aa765ae6e2a6e7212164640",
        63_192,
        "text/csv",
        "inventory_only",
    ),
}

_FDA_HEADERS = (
    "row",
    "ticker",
    "name",
    "catalyst_price_movement",
    "price_at_catalyst_date",
    "drug",
    "indication",
    "stage",
    "catalyst_date",
    "catalyst",
    "conference",
    "company_url",
    "catalyst_url",
)
_DATE_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})(?:\s+ET)?$", re.IGNORECASE)
_AUTHORITY = {
    "classification": "licensed_historical_context",
    "decision_authority": False,
    "allowed_uses": ["display", "context", "explain"],
    "forbidden_uses": [
        "originate_signal",
        "rank_security",
        "select_security",
        "size_position",
        "gate_decision",
        "execute_trade",
        "raise_authority",
    ],
}


def canonical_json_bytes(value: object) -> bytes:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")


def _clean(value: object, *, maximum: int) -> str | None:
    if value is None:
        return None
    rendered = unicodedata.normalize("NFKC", str(value)).replace("\ufeff", "").replace("\u00a0", " ")
    rendered = " ".join(rendered.split()).strip()
    if not rendered:
        return None
    return rendered[:maximum]


def _clean_market(value: object) -> str | None:
    rendered = _clean(value, maximum=64)
    if rendered is None or rendered.casefold() in {"-", "--", "—", "n/a", "na", "none", "null"}:
        return None
    return rendered


_PRIVATE_LOCATOR_VALUE = re.compile(r"(?:\b(?:https?|s3|r2|file)://\S+|\S*/raw/\S*)", re.IGNORECASE)


def _clean_description(value: object) -> tuple[str | None, bool]:
    rendered = _clean(value, maximum=2_400)
    if rendered is None:
        return None, False
    redacted = _PRIVATE_LOCATOR_VALUE.sub("[source link omitted]", rendered)
    return _clean(redacted, maximum=600), redacted != rendered


def admit_files(
    paths: Mapping[str, Path],
    *,
    specs: Mapping[str, InputSpec] = AUTHORIZED_INPUTS,
) -> tuple[AdmittedInput, ...]:
    if set(paths) != set(specs):
        raise SnapshotError("AUTHORIZED_INPUT_SET_INVALID")
    admitted: list[AdmittedInput] = []
    for input_id in specs:
        spec = specs[input_id]
        path = Path(paths[input_id])
        if not path.is_file() or path.is_symlink():
            raise SnapshotError("AUTHORIZED_INPUT_MISSING")
        data = path.read_bytes()
        if len(data) != spec.byte_count or sha256(data).hexdigest() != spec.sha256:
            raise SnapshotError("AUTHORIZED_INPUT_HASH_MISMATCH")
        admitted.append(AdmittedInput(spec=spec, path=path, data=data))
    return tuple(admitted)


def build_snapshot_manifest(
    admitted: Sequence[AdmittedInput],
    *,
    families: Sequence[Mapping[str, object]],
    observed_at: str,
    source_published_at: str | None = None,
) -> dict[str, Any]:
    if len(admitted) != 5 or {item.spec.input_id for item in admitted} != set(AUTHORIZED_INPUTS):
        raise SnapshotError("AUTHORIZED_INPUT_SET_INVALID")
    manifest: dict[str, Any] = {
        "contract_id": "biocatalyst_jv_snapshot_manifest.v1",
        "schema_version": "1.0.0",
        "source_id": "biopharmcatalyst_jv_snapshot",
        "license_class": "licensed_finite_snapshot",
        "capture": {
            "observed_at": observed_at,
            "source_published_at": source_published_at,
            "source_published_at_state": "observed" if source_published_at else "unknown",
        },
        "inputs": [
            {
                "input_id": item.spec.input_id,
                "safe_name": item.spec.safe_name,
                "sha256": item.spec.sha256,
                "byte_count": item.spec.byte_count,
                "media_type": item.spec.media_type,
                "role": item.spec.role,
            }
            for item in sorted(admitted, key=lambda value: value.spec.input_id)
        ],
        "families": [dict(row) for row in families],
    }
    manifest["manifest_sha256"] = sha256(canonical_json_bytes(manifest)).hexdigest()
    return manifest


def _cell(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return _clean(value, maximum=16_384)
    if isinstance(value, float) and value == 0:
        return 0
    return value


def _worksheet_payloads(data: bytes) -> list[tuple[str, list[list[object]]]]:
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    except Exception:
        raise SnapshotError("WORKBOOK_INVALID") from None
    payloads: list[tuple[str, list[list[object]]]] = []
    try:
        for sheet in workbook.worksheets:
            rows = [[_cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            while rows and all(value is None for value in rows[-1]):
                rows.pop()
            width = max((len(row) for row in rows), default=0)
            for row in rows:
                row.extend([None] * (width - len(row)))
            payloads.append((sheet.title, rows))
    finally:
        workbook.close()
    return payloads


def classify_workbook_bytes(older: bytes, newer: bytes) -> str:
    old = _worksheet_payloads(older)
    new = _worksheet_payloads(newer)
    old_names = [name for name, _ in old]
    new_names = [name for name, _ in new]
    if not set(old_names).issubset(new_names):
        return "DISTINCT_CAPTURE"
    new_map = dict(new)
    if any(canonical_json_bytes(rows) != canonical_json_bytes(new_map[name]) for name, rows in old):
        return "COMMON_SHEET_CONTENT_CHANGED"
    if old_names == new_names:
        return "UNRESOLVED"
    return "ADDITIVE_SHEET_EXPORT_IDENTICAL_COMMON_CONTENT"


def repair_historical_fda(
    raw: bytes,
    *,
    expected_rows: int | None = None,
    expected_shifted: int | None = None,
) -> HistoricalFdaRepair:
    try:
        reader = csv.reader(StringIO(raw.decode("utf-8-sig", errors="strict"), newline=""))
        header = tuple(next(reader))
    except Exception:
        raise SnapshotError("HISTORICAL_FDA_INVALID") from None
    if header != _FDA_HEADERS:
        raise SnapshotError("HISTORICAL_FDA_HEADER_INVALID")
    repaired: list[dict[str, str]] = []
    shifted = 0
    for values in reader:
        if not values or not any(values):
            continue
        repair = "none"
        if not values[0].strip().isdigit():
            values = [""] + values
            shifted += 1
            repair = "missing_row_index_unshifted"
        if len(values) != len(_FDA_HEADERS):
            raise SnapshotError("HISTORICAL_FDA_ROW_INVALID")
        row = dict(zip(_FDA_HEADERS, values, strict=True))
        row["_repair"] = repair
        repaired.append(row)
    if expected_rows is not None and len(repaired) != expected_rows:
        raise SnapshotError("HISTORICAL_FDA_ROW_COUNT_MISMATCH")
    if expected_shifted is not None and shifted != expected_shifted:
        raise SnapshotError("HISTORICAL_FDA_REPAIR_COUNT_MISMATCH")
    return HistoricalFdaRepair(tuple(repaired), len(repaired), shifted, sha256(raw).hexdigest())


def _event_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    rendered = _clean(value, maximum=64)
    if rendered is None:
        return None
    match = _DATE_RE.fullmatch(rendered)
    if match:
        try:
            return date(int(match.group(3)), int(match.group(2)), int(match.group(1))).isoformat()
        except ValueError:
            return None
    try:
        return date.fromisoformat(rendered).isoformat()
    except ValueError:
        return None


def _default_identity(_ticker: str | None, _event_date: date) -> dict[str, object]:
    return {
        "resolution_state": "unresolved",
        "security_id": None,
        "issuer_id": None,
        "resolution_basis": "none",
        "issuer_relationship_state": "unavailable",
    }


def _make_event(
    *,
    source_family: str,
    source_ordinal: int,
    ticker: object,
    name: object,
    event_date: object,
    event_family: str,
    stage: object,
    description: object,
    asset_kind: str,
    asset_label: object,
    indication: object,
    price: object,
    movement: object,
    repair: str,
    unsafe_fields: Sequence[str],
    observed_at: str,
    source_published_at: str | None,
    identity_resolver: Callable[[str | None, date], Mapping[str, object]],
) -> dict[str, Any] | None:
    normalized_date = _event_date(event_date)
    if normalized_date is None:
        return None
    ticker_value = _clean(ticker, maximum=32)
    identity = dict(identity_resolver(ticker_value, date.fromisoformat(normalized_date)))
    description_value, description_redacted = _clean_description(description)
    unsafe = set(unsafe_fields)
    if description_redacted:
        unsafe.add("licensed_description_urls_redacted")
    semantic = {
        "source_family": source_family,
        "ticker": ticker_value,
        "name": _clean(name, maximum=240),
        "date": normalized_date,
        "family": event_family,
        "stage": _clean(stage, maximum=160),
        "description": description_value,
        "asset_kind": asset_kind,
        "asset_label": _clean(asset_label, maximum=240),
        "indication": _clean(indication, maximum=320),
        "price": _clean_market(price),
        "movement": _clean_market(movement),
        "repair": repair,
    }
    event_id = "bpcjv_event_" + sha256(canonical_json_bytes(semantic)).hexdigest()[:24]
    return {
        "contract_id": "biocatalyst_historical_event_record.v1",
        "schema_version": "1.0.0",
        "event_id": event_id,
        "source": {
            "provider": "BioPharmCatalyst",
            "source_id": "biopharmcatalyst_jv_snapshot",
            "license_class": "licensed_finite_snapshot",
            "family": source_family,
            "source_ordinal": source_ordinal,
            "capture_observed_at": observed_at,
            "source_published_at": source_published_at,
            "source_published_at_state": "observed" if source_published_at else "unknown",
        },
        "company": {
            "ticker_evidence": ticker_value,
            "name_evidence": semantic["name"],
            "resolution_state": identity.get("resolution_state", "unresolved"),
            "security_id": identity.get("security_id"),
            "issuer_id": identity.get("issuer_id"),
            "resolution_basis": identity.get("resolution_basis", "none"),
            "issuer_relationship_state": identity.get("issuer_relationship_state", "unavailable"),
        },
        "event": {
            "date": normalized_date,
            "date_precision": "day",
            "family": event_family,
            "stage": semantic["stage"],
            "description": semantic["description"],
            "source_available_at": None,
            "observed_at": observed_at,
        },
        "asset": {"kind": asset_kind, "label": semantic["asset_label"], "indication": semantic["indication"]},
        "historical_market": {"price_at_event": semantic["price"], "price_movement": semantic["movement"]},
        "normalization": {"state": "deterministic", "repair": repair},
        "unsafe_fields": sorted(unsafe),
        "authority": dict(_AUTHORITY),
    }


def _rows_by_sheet(workbook_bytes: bytes) -> dict[str, tuple[list[str], list[list[object]]]]:
    output: dict[str, tuple[list[str], list[list[object]]]] = {}
    for title, rows in _worksheet_payloads(workbook_bytes):
        if not rows:
            continue
        headers = [str(value or "").strip() for value in rows[0]]
        output[title] = (headers, rows[1:])
    return output


def _record(headers: Sequence[str], values: Sequence[object]) -> dict[str, object]:
    return {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}


def normalize_corpus(
    workbook_bytes: bytes,
    historical_fda_bytes: bytes,
    *,
    observed_at: str,
    source_published_at: str | None = None,
    expected_fda_rows: int | None = 15_700,
    expected_fda_shifted: int | None = 4_404,
    identity_resolver: Callable[[str | None, date], Mapping[str, object]] | None = None,
) -> NormalizedCorpus:
    try:
        capture_day = datetime.fromisoformat(observed_at.replace("Z", "+00:00")).date()
    except ValueError:
        raise SnapshotError("CAPTURE_CLOCK_INVALID") from None
    resolver = identity_resolver or _default_identity
    repaired = repair_historical_fda(
        historical_fda_bytes,
        expected_rows=expected_fda_rows,
        expected_shifted=expected_fda_shifted,
    )
    events: list[dict[str, Any]] = []
    for ordinal, row in enumerate(repaired.rows, start=1):
        event = _make_event(
            source_family="historical_fda",
            source_ordinal=ordinal,
            ticker=row["ticker"],
            name=row["name"],
            event_date=row["catalyst_date"],
            event_family="regulatory",
            stage=row["stage"],
            description=row["catalyst"],
            asset_kind="drug",
            asset_label=row["drug"],
            indication=row["indication"],
            price=row["price_at_catalyst_date"],
            movement=row["catalyst_price_movement"],
            repair=row["_repair"],
            unsafe_fields=("licensed_source_links_omitted", "capture_only_overlays_unavailable"),
            observed_at=observed_at,
            source_published_at=source_published_at,
            identity_resolver=resolver,
        )
        if event is not None:
            events.append(event)

    sheets = _rows_by_sheet(workbook_bytes)
    device_history = sheets.get("Device History")
    if device_history:
        headers, rows = device_history
        for ordinal, values in enumerate(rows, start=2):
            row = _record(headers, values)
            event = _make_event(
                source_family="device_history",
                source_ordinal=ordinal,
                ticker=row.get("Ticker"),
                name=row.get("Name"),
                event_date=row.get("Catalyst Date"),
                event_family="device",
                stage=row.get("Device Stage"),
                description=row.get("Catalyst"),
                asset_kind="device",
                asset_label=row.get("Device"),
                indication=row.get("Indication"),
                price=row.get("Price at Catalyst Date"),
                movement=row.get("Catalyst Price Movement"),
                repair="none",
                unsafe_fields=("capture_only_overlays_unavailable",),
                observed_at=observed_at,
                source_published_at=source_published_at,
                identity_resolver=resolver,
            )
            if event is not None:
                events.append(event)

    device_pipeline = sheets.get("Device Pipeline")
    if device_pipeline:
        headers, rows = device_pipeline
        for ordinal, values in enumerate(rows, start=2):
            row = _record(headers, values)
            event = _make_event(
                source_family="device_pipeline_history",
                source_ordinal=ordinal,
                ticker=row.get("Ticker"),
                name=row.get("Name"),
                event_date=row.get("Catalyst Date"),
                event_family="device",
                stage=row.get("Stage"),
                description=row.get("Catalyst"),
                asset_kind="device",
                asset_label=row.get("Device"),
                indication=row.get("Indication"),
                price=None,
                movement=None,
                repair="none",
                unsafe_fields=("capture_only_market_overlays_omitted", "capture_only_options_omitted"),
                observed_at=observed_at,
                source_published_at=source_published_at,
                identity_resolver=resolver,
            )
            if event is not None:
                events.append(event)

    # This projection is historical context, not the workbook's forward
    # catalyst calendar. Future-due rows remain admitted/inventoried in the
    # family denominator but cannot masquerade as events that already happened.
    events = [event for event in events if date.fromisoformat(event["event"]["date"]) <= capture_day]

    unique_events: dict[str, dict[str, Any]] = {}
    duplicates_collapsed = 0
    for event in events:
        event_id = event["event_id"]
        existing = unique_events.get(event_id)
        if existing is None:
            unique_events[event_id] = event
            continue
        existing_compare = json.loads(canonical_json_bytes(existing))
        event_compare = json.loads(canonical_json_bytes(event))
        existing_compare["source"]["source_ordinal"] = 0
        event_compare["source"]["source_ordinal"] = 0
        if canonical_json_bytes(existing_compare) != canonical_json_bytes(event_compare):
            raise SnapshotError("DUPLICATE_NORMALIZED_EVENT_DIVERGENT")
        duplicates_collapsed += 1
    events = list(unique_events.values())
    events.sort(key=lambda item: (item["event"]["date"], item["event_id"]), reverse=True)
    jsonl = b"".join(canonical_json_bytes(event) + b"\n" for event in events)
    family_counts: dict[str, int] = {}
    resolved = 0
    for event in events:
        family = event["source"]["family"]
        family_counts[family] = family_counts.get(family, 0) + 1
        resolved += int(event["company"]["resolution_state"] == "resolved")
    family_source_rows = {
        "historical_fda": repaired.row_count,
        "device_history": len(device_history[1]) if device_history else 0,
        "device_pipeline_history": len(device_pipeline[1]) if device_pipeline else 0,
    }
    coverage = {
        "state": "partial",
        "source_rows": sum(family_source_rows.values()),
        "normalized_rows": len(events),
        "identity_resolved": resolved,
        "identity_unresolved": len(events) - resolved,
        "duplicates_collapsed": duplicates_collapsed,
        "families": dict(sorted(family_counts.items())),
        "family_source_rows": family_source_rows,
    }
    return NormalizedCorpus(tuple(events), jsonl, coverage, repaired)
